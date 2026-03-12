import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import math

# ----------------------------------------------------------------
# PAGE CONFIG
# ----------------------------------------------------------------
st.set_page_config(
    page_title="MaintEdge - DMC Calculator | Deutsche Aircraft",
    page_icon="https://img.icons8.com/fluency/48/airplane-mode-on.png",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ----------------------------------------------------------------
# DO 328-100 MAINTENANCE DATA (baseline from Excel)
# ----------------------------------------------------------------
DO328_100_DATA = [
    # Airframe Checks (updated LC material to 14.08)
    {"inspection": "Engine Oil Level Check", "int1": None, "param1": None,     "int2": 24,    "param2": "FH",      "mh": 0.5,   "mat": 37.55,      "category": "Airframe Checks"},
    {"inspection": "LC",              "int1": 15,   "param1": "Days",   "int2": 60,    "param2": "FH",      "mh": 8,     "mat": 14.08,      "category": "Airframe Checks"},
    {"inspection": "A1",              "int1": 7.5,  "param1": "Months", "int2": 500,   "param2": "FH",      "mh": 80,    "mat": 91.23,      "category": "Airframe Checks"},
    {"inspection": "A2",              "int1": 15,   "param1": "Months", "int2": 1000,  "param2": "FH",      "mh": 50,    "mat": 10847.72,   "category": "Airframe Checks"},
    {"inspection": "A3",              "int1": 15,   "param1": "Months", "int2": 1500,  "param2": "FH",      "mh": 20,    "mat": 11383.37,   "category": "Airframe Checks"},
    {"inspection": "A4",              "int1": 30,   "param1": "Months", "int2": 2000,  "param2": "FH",      "mh": 18,    "mat": 136.86,     "category": "Airframe Checks"},
    {"inspection": "A5",              "int1": 30,   "param1": "Months", "int2": 2500,  "param2": "FH",      "mh": 75,    "mat": 566.35,     "category": "Airframe Checks"},
    {"inspection": "C1",              "int1": 30,   "param1": "Months", "int2": 5000,  "param2": "FH",      "mh": 360,   "mat": 8682.17,    "category": "Airframe Checks"},
    {"inspection": "C2",              "int1": 60,   "param1": "Months", "int2": 10000, "param2": "FH",      "mh": 270,   "mat": 1824.69,    "category": "Airframe Checks"},
    {"inspection": "C3",              "int1": 90,   "param1": "Months", "int2": 15000, "param2": "FH",      "mh": 30,    "mat": 182.47,     "category": "Airframe Checks"},
    {"inspection": "C4",              "int1": 120,  "param1": "Months", "int2": 20000, "param2": "FH",      "mh": 140,   "mat": 2982.48,    "category": "Airframe Checks"},
    # APU Inspections (unchanged)
    {"inspection": "APU500",          "int1": None, "param1": None,     "int2": 500,   "param2": "APU Hrs", "mh": 2.5,   "mat": 217.62,     "category": "APU Inspections"},
    {"inspection": "APU800",          "int1": None, "param1": None,     "int2": 8000,  "param2": "APU Hrs", "mh": 8,     "mat": 249.02,     "category": "APU Inspections"},
    {"inspection": "APU1000",         "int1": None, "param1": None,     "int2": 1000,  "param2": "APU Hrs", "mh": 11,    "mat": 0,          "category": "APU Inspections"},
    {"inspection": "APU5000",         "int1": None, "param1": None,     "int2": 5000,  "param2": "APU Hrs", "mh": 10,    "mat": 255.00,     "category": "APU Inspections"},
    # FH-Based Tasks (unchanged)
    {"inspection": "FH1000",          "int1": None, "param1": None,     "int2": 1000,  "param2": "FH",      "mh": 10.5,  "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH1200",          "int1": None, "param1": None,     "int2": 1200,  "param2": "FH",      "mh": 2,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH2000",          "int1": None, "param1": None,     "int2": 2000,  "param2": "FH",      "mh": 4,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH4000",          "int1": None, "param1": None,     "int2": 4000,  "param2": "FH",      "mh": 11.5,  "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH4000/60M",      "int1": 60,   "param1": "Months", "int2": 4000,  "param2": "FH",      "mh": 5,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH8000",          "int1": None, "param1": None,     "int2": 8000,  "param2": "FH",      "mh": 3,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "Weight & Balance","int1": 48,   "param1": "Months", "int2": None,  "param2": None,      "mh": 23,    "mat": 0,          "category": "FH-Based Tasks"},
    # Fatigue Damage (unchanged)
    {"inspection": "FD-0101",         "int1": None, "param1": None,     "int2": 1000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0202",         "int1": None, "param1": None,     "int2": 2000,  "param2": "FC",      "mh": 6,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0303",         "int1": None, "param1": None,     "int2": 3000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0404",         "int1": None, "param1": None,     "int2": 4000,  "param2": "FC",      "mh": 4,     "mat": 47.76,      "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0602",         "int1": None, "param1": None,     "int2": 6000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0606",         "int1": None, "param1": None,     "int2": 6000,  "param2": "FC",      "mh": 8,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0808",         "int1": None, "param1": None,     "int2": 8000,  "param2": "FC",      "mh": 30,    "mat": 278.52,     "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0808A",        "int1": None, "param1": None,     "int2": 8000,  "param2": "FC",      "mh": 18,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1000",         "int1": None, "param1": None,     "int2": 10000, "param2": "FC",      "mh": 8,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1212",         "int1": None, "param1": None,     "int2": 12000, "param2": "FC",      "mh": 9,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1616",         "int1": None, "param1": None,     "int2": 16000, "param2": "FC",      "mh": 12,    "mat": 189.03,     "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1616A",        "int1": None, "param1": None,     "int2": 16000, "param2": "FC",      "mh": 4,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2000",         "int1": None, "param1": None,     "int2": 20000, "param2": "FC",      "mh": 3,     "mat": 4839.81,    "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2208",         "int1": None, "param1": None,     "int2": 22000, "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2502",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 6,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2503",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2504",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 14,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2506",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 45,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2508",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 105,   "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2508A",        "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 12,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2510",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 24,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2512",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 90,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2516",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 225,   "mat": 0,          "category": "Fatigue Damage (FD)"},
    # Corrosion Prevention (unchanged)
    {"inspection": "CP-2",            "int1": 24,   "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-2.5",          "int1": 30,   "param1": "Months", "int2": None,  "param2": None,      "mh": 20,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-4",            "int1": 48,   "param1": "Months", "int2": None,  "param2": None,      "mh": 10,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-5",            "int1": 60,   "param1": "Months", "int2": None,  "param2": None,      "mh": 35,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-10",           "int1": 96,   "param1": "Months", "int2": None,  "param2": None,      "mh": 105,   "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-10/5",         "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Corrosion Prevention (CP)"},
    # Structural Sampling (SSI-10/5 material corrected to 0)
    {"inspection": "SSI-2",           "int1": 24,   "param1": "Months", "int2": None,  "param2": None,      "mh": 15,    "mat": 0,          "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-2.5",         "int1": 30,   "param1": "Months", "int2": None,  "param2": None,      "mh": 10,    "mat": 314.87,     "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-5",           "int1": 60,   "param1": "Months", "int2": None,  "param2": None,      "mh": 80,    "mat": 1096.10,    "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-10",          "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 220,   "mat": 2991.68,    "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-10/5",        "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Structural Sampling (SSI)"},
    # Heavy Components (updated prices from new spreadsheet, quantities applied)
    {"inspection": "Propeller Change (2EA)",  "int1": None, "param1": None,     "int2": 6000,  "param2": "FH",  "mh": 80,    "mat": 350000.00,  "category": "Propellers"},
    {"inspection": "Engine Change (2EA)",     "int1": None, "param1": None,     "int2": 8000,  "param2": "FH",  "mh": 360,   "mat": 2400000.00, "category": "Engines"},
    {"inspection": "Landing Gear Overhaul",   "int1": 144,  "param1": "Months", "int2": 22000, "param2": "FC",  "mh": 85,    "mat": 850000.00,  "category": "Landing Gear"},
    {"inspection": "Brakes (4EA)",            "int1": None, "param1": None,     "int2": 3000,  "param2": "FC",  "mh": 8,     "mat": 80000.00,   "category": "Landing Gear"},
    {"inspection": "NLG Tires (2EA)",         "int1": None, "param1": None,     "int2": 250,   "param2": "FC",  "mh": 4,     "mat": 1746.72,    "category": "Landing Gear"},
    {"inspection": "MLG Tires (4EA)",         "int1": None, "param1": None,     "int2": 150,   "param2": "FC",  "mh": 16,    "mat": 7600.00,    "category": "Landing Gear"},
    {"inspection": "APU Overhaul",            "int1": None, "param1": None,     "int2": 8000,  "param2": "FC",  "mh": 16,    "mat": 123000.00,  "category": "APU"},
    # Time Controlled Components (MOD 10 baseline; MOD 20/30 factors applied via MOD_FACTORS)
    {"inspection": "Time Controlled Items",   "int1": None, "param1": None,     "int2": 2000,  "param2": "FH",  "mh": 40,    "mat": 75000.00,   "category": "Time Controlled Components"},
]


# ----------------------------------------------------------------
# DO 328-300 (JET) MAINTENANCE DATA (from Excel: DMC calculation d328 jet)
# Engine: PW306B x2 | No propellers
# ----------------------------------------------------------------
D328_300_JET_DATA = [
    # Airframe Checks
    {"inspection": "Engine Oil Level Check", "int1": None, "param1": None,     "int2": 24,    "param2": "FH",      "mh": 0.5,   "mat": 37.55,      "category": "Airframe Checks"},
    {"inspection": "LC",              "int1": 15,   "param1": "Days",   "int2": 60,    "param2": "FH",      "mh": 8,     "mat": 14.08,      "category": "Airframe Checks"},
    {"inspection": "A1",              "int1": 7.5,  "param1": "Months", "int2": 500,   "param2": "FH",      "mh": 60,    "mat": 91.23,      "category": "Airframe Checks"},
    {"inspection": "A2",              "int1": 15,   "param1": "Months", "int2": 1000,  "param2": "FH",      "mh": 20,    "mat": 1593.58,    "category": "Airframe Checks"},
    {"inspection": "A3",              "int1": 15,   "param1": "Months", "int2": 1500,  "param2": "FH",      "mh": 19,    "mat": 6655.95,    "category": "Airframe Checks"},
    {"inspection": "A4",              "int1": 30,   "param1": "Months", "int2": 2000,  "param2": "FH",      "mh": 14,    "mat": 136.86,     "category": "Airframe Checks"},
    {"inspection": "A5 (15M)",        "int1": 15,   "param1": "Months", "int2": None,  "param2": None,      "mh": 14,    "mat": 0,          "category": "Airframe Checks"},
    {"inspection": "A5 (30M/2500FH)", "int1": 30,   "param1": "Months", "int2": 2500,  "param2": "FH",      "mh": 50,    "mat": 566.35,     "category": "Airframe Checks"},
    {"inspection": "C1",              "int1": 30,   "param1": "Months", "int2": 5000,  "param2": "FH",      "mh": 354,   "mat": 9582.96,    "category": "Airframe Checks"},
    {"inspection": "C2",              "int1": 60,   "param1": "Months", "int2": 10000, "param2": "FH",      "mh": 309,   "mat": 3969.98,    "category": "Airframe Checks"},
    {"inspection": "C3",              "int1": 90,   "param1": "Months", "int2": 15000, "param2": "FH",      "mh": 35,    "mat": 182.47,     "category": "Airframe Checks"},
    {"inspection": "C4",              "int1": 120,  "param1": "Months", "int2": 20000, "param2": "FH",      "mh": 143,   "mat": 17180.50,   "category": "Airframe Checks"},
    # APU Inspections
    {"inspection": "APU500",          "int1": None, "param1": None,     "int2": 500,   "param2": "APU Hrs", "mh": 2.5,   "mat": 217.62,     "category": "APU Inspections"},
    {"inspection": "APU1000",         "int1": None, "param1": None,     "int2": 1000,  "param2": "APU Hrs", "mh": 11,    "mat": 0,          "category": "APU Inspections"},
    {"inspection": "APU5000",         "int1": None, "param1": None,     "int2": 5000,  "param2": "APU Hrs", "mh": 10,    "mat": 255.00,     "category": "APU Inspections"},
    # FH-Based Tasks
    {"inspection": "Compressor Wash", "int1": None, "param1": None,     "int2": 1000,  "param2": "FH",      "mh": 10,    "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH1000",          "int1": None, "param1": None,     "int2": 1000,  "param2": "FH",      "mh": 10.5,  "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH1200",          "int1": None, "param1": None,     "int2": 1200,  "param2": "FH",      "mh": 2,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH2000",          "int1": None, "param1": None,     "int2": 2000,  "param2": "FH",      "mh": 4,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH4000",          "int1": None, "param1": None,     "int2": 4000,  "param2": "FH",      "mh": 11.5,  "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH4000/60M",      "int1": 60,   "param1": "Months", "int2": 4000,  "param2": "FH",      "mh": 5,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH8000",          "int1": None, "param1": None,     "int2": 8000,  "param2": "FH",      "mh": 3,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "Weight & Balance","int1": 48,   "param1": "Months", "int2": None,  "param2": None,      "mh": 23,    "mat": 0,          "category": "FH-Based Tasks"},
    # Fatigue Damage (FD)
    {"inspection": "FD-0101",         "int1": None, "param1": None,     "int2": 1000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0202",         "int1": None, "param1": None,     "int2": 2000,  "param2": "FC",      "mh": 6,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0303",         "int1": None, "param1": None,     "int2": 3000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0404",         "int1": None, "param1": None,     "int2": 4000,  "param2": "FC",      "mh": 3,     "mat": 47.76,      "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0602",         "int1": None, "param1": None,     "int2": 6000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0606",         "int1": None, "param1": None,     "int2": 6000,  "param2": "FC",      "mh": 8,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0808",         "int1": None, "param1": None,     "int2": 8000,  "param2": "FC",      "mh": 30,    "mat": 278.52,     "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0808A",        "int1": None, "param1": None,     "int2": 8000,  "param2": "FC",      "mh": 8,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1000",         "int1": None, "param1": None,     "int2": 10000, "param2": "FC",      "mh": 2,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1212",         "int1": None, "param1": None,     "int2": 12000, "param2": "FC",      "mh": 9,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1616",         "int1": None, "param1": None,     "int2": 16000, "param2": "FC",      "mh": 12,    "mat": 189.03,     "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1616A",        "int1": None, "param1": None,     "int2": 16000, "param2": "FC",      "mh": 1.5,   "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2000",         "int1": None, "param1": None,     "int2": 20000, "param2": "FC",      "mh": 3,     "mat": 4839.81,    "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2208",         "int1": None, "param1": None,     "int2": 22000, "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2502",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 6,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2503",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2504",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 9,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2506",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 45,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2508",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 105,   "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2512",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 90,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2516",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 225,   "mat": 0,          "category": "Fatigue Damage (FD)"},
    # Corrosion Prevention (CP)
    {"inspection": "CP-2",            "int1": 24,   "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-2.5",          "int1": 30,   "param1": "Months", "int2": None,  "param2": None,      "mh": 20,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-4",            "int1": 48,   "param1": "Months", "int2": None,  "param2": None,      "mh": 10,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-5",            "int1": 60,   "param1": "Months", "int2": None,  "param2": None,      "mh": 35,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-10",           "int1": 96,   "param1": "Months", "int2": None,  "param2": None,      "mh": 105,   "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-10/5",         "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Corrosion Prevention (CP)"},
    # Structural Sampling (SSI)
    {"inspection": "SSI-2",           "int1": 24,   "param1": "Months", "int2": None,  "param2": None,      "mh": 15,    "mat": 0,          "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-2.5",         "int1": 30,   "param1": "Months", "int2": None,  "param2": None,      "mh": 10,    "mat": 314.87,     "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-5",           "int1": 60,   "param1": "Months", "int2": None,  "param2": None,      "mh": 80,    "mat": 1096.10,    "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-10",          "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 220,   "mat": 2991.68,    "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-10/5",        "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Structural Sampling (SSI)"},
    # Heavy Components (quantities applied per notes in Excel)
    {"inspection": "Engine Change (2EA)",   "int1": None, "param1": None,     "int2": 8000,  "param2": "FH",      "mh": 360,   "mat": 3304050.00, "category": "Engines"},
    {"inspection": "APU Overhaul",          "int1": None, "param1": None,     "int2": 6000,  "param2": "APU Hrs", "mh": 16,    "mat": 123000.00,  "category": "APU"},
    {"inspection": "Landing Gear Overhaul", "int1": 144,  "param1": "Months", "int2": 22000, "param2": "FC",      "mh": 85,    "mat": 850000.00,  "category": "Landing Gear"},
    {"inspection": "Brakes (4EA)",          "int1": None, "param1": None,     "int2": 3000,  "param2": "FC",      "mh": 8,     "mat": 80000.00,   "category": "Landing Gear"},
    {"inspection": "NLG Tires (2EA)",       "int1": None, "param1": None,     "int2": 250,   "param2": "FC",      "mh": 4,     "mat": 1746.72,    "category": "Landing Gear"},
    {"inspection": "MLG Tires (4EA)",       "int1": None, "param1": None,     "int2": 150,   "param2": "FC",      "mh": 16,    "mat": 7600.00,    "category": "Landing Gear"},
    # Time Controlled Components
    {"inspection": "Time Controlled Items", "int1": None, "param1": None,     "int2": 2000,  "param2": "FH",      "mh": 40,    "mat": 75000.00,   "category": "Time Controlled Components"},
]


# ----------------------------------------------------------------
# DO 328eco MAINTENANCE DATA
# Engine: PW127XT-S x2 via PWC FMP (pay-per-hour) -- NO APU
# Propeller: 8,000 FH / 84 months | LG overhaul: $530,000
# Source: DMC calculation D328eco.xlsx + PWC FMP PDF (Sept 2023)
# ----------------------------------------------------------------
D328_ECO_DATA = [
    # Airframe Checks
    {"inspection": "Engine Oil Level Check", "int1": None, "param1": None,     "int2": 24,    "param2": "FH",      "mh": 0.5,   "mat": 37.55,      "category": "Airframe Checks"},
    {"inspection": "LC",              "int1": 15,   "param1": "Days",   "int2": 60,    "param2": "FH",      "mh": 8,     "mat": 14.08,      "category": "Airframe Checks"},
    {"inspection": "A1",              "int1": 7.5,  "param1": "Months", "int2": 500,   "param2": "FH",      "mh": 80,    "mat": 91.23,      "category": "Airframe Checks"},
    {"inspection": "A2",              "int1": 15,   "param1": "Months", "int2": 1000,  "param2": "FH",      "mh": 50,    "mat": 10847.72,   "category": "Airframe Checks"},
    {"inspection": "A3",              "int1": 15,   "param1": "Months", "int2": 1500,  "param2": "FH",      "mh": 20,    "mat": 11383.37,   "category": "Airframe Checks"},
    {"inspection": "A4",              "int1": 30,   "param1": "Months", "int2": 2000,  "param2": "FH",      "mh": 18,    "mat": 136.86,     "category": "Airframe Checks"},
    {"inspection": "A5",              "int1": 30,   "param1": "Months", "int2": 2500,  "param2": "FH",      "mh": 75,    "mat": 566.35,     "category": "Airframe Checks"},
    {"inspection": "C1",              "int1": 30,   "param1": "Months", "int2": 5000,  "param2": "FH",      "mh": 360,   "mat": 8682.17,    "category": "Airframe Checks"},
    {"inspection": "C2",              "int1": 60,   "param1": "Months", "int2": 10000, "param2": "FH",      "mh": 270,   "mat": 1824.69,    "category": "Airframe Checks"},
    {"inspection": "C3",              "int1": 90,   "param1": "Months", "int2": 15000, "param2": "FH",      "mh": 30,    "mat": 182.47,     "category": "Airframe Checks"},
    {"inspection": "C4",              "int1": 120,  "param1": "Months", "int2": 20000, "param2": "FH",      "mh": 140,   "mat": 2982.48,    "category": "Airframe Checks"},
    # FH-Based Tasks (no APU tasks -- D328eco has no APU)
    {"inspection": "FH1000",          "int1": None, "param1": None,     "int2": 1000,  "param2": "FH",      "mh": 10.5,  "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH1200",          "int1": None, "param1": None,     "int2": 1200,  "param2": "FH",      "mh": 2,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH2000",          "int1": None, "param1": None,     "int2": 2000,  "param2": "FH",      "mh": 4,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH4000",          "int1": None, "param1": None,     "int2": 4000,  "param2": "FH",      "mh": 11.5,  "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH4000/60M",      "int1": 60,   "param1": "Months", "int2": 4000,  "param2": "FH",      "mh": 5,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "FH8000",          "int1": None, "param1": None,     "int2": 8000,  "param2": "FH",      "mh": 3,     "mat": 0,          "category": "FH-Based Tasks"},
    {"inspection": "Weight & Balance","int1": 48,   "param1": "Months", "int2": None,  "param2": None,      "mh": 23,    "mat": 0,          "category": "FH-Based Tasks"},
    # Fatigue Damage (FD)
    {"inspection": "FD-0101",         "int1": None, "param1": None,     "int2": 1000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0202",         "int1": None, "param1": None,     "int2": 2000,  "param2": "FC",      "mh": 6,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0303",         "int1": None, "param1": None,     "int2": 3000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0404",         "int1": None, "param1": None,     "int2": 4000,  "param2": "FC",      "mh": 4,     "mat": 47.76,      "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0602",         "int1": None, "param1": None,     "int2": 6000,  "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0606",         "int1": None, "param1": None,     "int2": 6000,  "param2": "FC",      "mh": 8,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0808",         "int1": None, "param1": None,     "int2": 8000,  "param2": "FC",      "mh": 30,    "mat": 278.52,     "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-0808A",        "int1": None, "param1": None,     "int2": 8000,  "param2": "FC",      "mh": 18,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1000",         "int1": None, "param1": None,     "int2": 10000, "param2": "FC",      "mh": 8,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1212",         "int1": None, "param1": None,     "int2": 12000, "param2": "FC",      "mh": 9,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1616",         "int1": None, "param1": None,     "int2": 16000, "param2": "FC",      "mh": 12,    "mat": 189.03,     "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-1616A",        "int1": None, "param1": None,     "int2": 16000, "param2": "FC",      "mh": 4,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2000",         "int1": None, "param1": None,     "int2": 20000, "param2": "FC",      "mh": 3,     "mat": 4839.81,    "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2208",         "int1": None, "param1": None,     "int2": 22000, "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2502",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 6,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2503",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 3,     "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2504",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 14,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2506",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 45,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2508",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 105,   "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2508A",        "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 12,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2510",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 24,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2512",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 90,    "mat": 0,          "category": "Fatigue Damage (FD)"},
    {"inspection": "FD-2516",         "int1": None, "param1": None,     "int2": 25000, "param2": "FC",      "mh": 225,   "mat": 0,          "category": "Fatigue Damage (FD)"},
    # Corrosion Prevention (CP)
    {"inspection": "CP-2",            "int1": 24,   "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-2.5",          "int1": 30,   "param1": "Months", "int2": None,  "param2": None,      "mh": 20,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-4",            "int1": 48,   "param1": "Months", "int2": None,  "param2": None,      "mh": 10,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-5",            "int1": 60,   "param1": "Months", "int2": None,  "param2": None,      "mh": 35,    "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-10",           "int1": 96,   "param1": "Months", "int2": None,  "param2": None,      "mh": 105,   "mat": 0,          "category": "Corrosion Prevention (CP)"},
    {"inspection": "CP-10/5",         "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Corrosion Prevention (CP)"},
    # Structural Sampling (SSI)
    {"inspection": "SSI-2",           "int1": 24,   "param1": "Months", "int2": None,  "param2": None,      "mh": 15,    "mat": 0,          "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-2.5",         "int1": 30,   "param1": "Months", "int2": None,  "param2": None,      "mh": 10,    "mat": 314.87,     "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-5",           "int1": 60,   "param1": "Months", "int2": None,  "param2": None,      "mh": 80,    "mat": 1096.10,    "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-10",          "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 220,   "mat": 2991.68,    "category": "Structural Sampling (SSI)"},
    {"inspection": "SSI-10/5",        "int1": 120,  "param1": "Months", "int2": None,  "param2": None,      "mh": 5,     "mat": 0,          "category": "Structural Sampling (SSI)"},
    # Time Controlled Components
    {"inspection": "Time Controlled Items",  "int1": None, "param1": None,    "int2": 2000, "param2": "FH", "mh": 40, "mat": 63000.00,   "category": "Time Controlled Components"},
    # Heavy Components (no APU, no engine overhaul -- engines via PWC FMP)
    {"inspection": "Propeller Change (2EA)", "int1": 84,  "param1": "Months", "int2": 8000, "param2": "FH", "mh": 80,  "mat": 358604.00,  "category": "Propellers"},
    {"inspection": "Landing Gear Overhaul", "int1": 144,  "param1": "Months", "int2": 22000,"param2": "FC", "mh": 85,  "mat": 530000.00,  "category": "Landing Gear"},
    {"inspection": "Brakes (4EA)",          "int1": None, "param1": None,     "int2": 3000, "param2": "FC", "mh": 8,   "mat": 80000.00,   "category": "Landing Gear"},
    {"inspection": "NLG Tires (2EA)",       "int1": None, "param1": None,     "int2": 250,  "param2": "FC", "mh": 4,   "mat": 1746.72,    "category": "Landing Gear"},
    {"inspection": "MLG Tires (4EA)",       "int1": None, "param1": None,     "int2": 150,  "param2": "FC", "mh": 16,  "mat": 7600.00,    "category": "Landing Gear"},
]

# ----------------------------------------------------------------
# PWC FMP ENGINE DMC -- PW127XT-S (D328eco)
# Source: "2023 09 DMC D328eco PW127XT-S (002) 2.pdf" (Pratt & Whitney Canada)
# IMPORTANT: PWC FMP rates are quoted PER ENGINE (industry standard).
# The D328eco has 2 x PW127XT-S engines, so the per-aircraft rate = 2 x table value.
# Services covered per engine: overhaul incl. shop labour & parts, HSI, LLP
#   replacement, engine BUR, SB cat 1-6, DPHM/ECTM, FAST connectivity
#   (incl. propeller balance monitoring, aircraft data services, engine data),
#   fuel nozzle refurbishment. NOTE: propeller overhaul is a separate line item.
# Escalation: 3.5% p.a. compound from Sept 2023 to March 2026 = 2.5 years
# USD to EUR conversion: 0.92 EUR/USD
# ----------------------------------------------------------------
_PWC_FMP_TABLE = [
    # (avg flight duration in minutes, USD/FH per SINGLE engine in $2023)
    (23,  243),
    (34,  193),
    (44,  164),
    (55,  146),
    (65,  135),
    (86,  135),
]
_PWC_ESCALATION_FACTOR = (1.035 ** 2.5)   # 3.5%/yr x 2.5 yrs = ~1.0899
_PWC_USD_TO_EUR = 0.92                     # USD to EUR
_PWC_ENGINE_COUNT = 2                      # D328eco has 2 engines


def get_pwc_engine_rate_eur(fh_per_year, fc_per_year):
    """
    Return the escalated EUR/FH engine DMC for PW127XT-S x2 (both engines) via PWC FMP.
    PWC rates are per single engine; multiplied by 2 for both engines per aircraft.
    Interpolates from the PWC pay-per-hour table based on average flight duration.
    Prices from PWC PDF (Sept 2023) escalated to March 2026 at 3.5%/yr, USD->EUR at 0.92.
    """
    if fc_per_year <= 0:
        return 0.0
    avg_min = (fh_per_year / fc_per_year) * 60.0  # average flight duration (minutes)

    table = _PWC_FMP_TABLE
    if avg_min <= table[0][0]:
        rate_usd_2023 = table[0][1]
    elif avg_min >= table[-1][0]:
        rate_usd_2023 = table[-1][1]
    else:
        rate_usd_2023 = table[0][1]
        for i in range(len(table) - 1):
            x0, x1 = table[i][0], table[i + 1][0]
            y0, y1 = table[i][1], table[i + 1][1]
            if x0 <= avg_min <= x1:
                frac = (avg_min - x0) / (x1 - x0)
                rate_usd_2023 = y0 + frac * (y1 - y0)
                break

    # Per engine rate escalated and converted, then x2 for both engines
    return rate_usd_2023 * _PWC_USD_TO_EUR * _PWC_ESCALATION_FACTOR * _PWC_ENGINE_COUNT


def get_aircraft_data(aircraft_type):
    """Return the maintenance data list for the selected aircraft type."""
    if "328-300" in aircraft_type or "Jet" in aircraft_type:
        return D328_300_JET_DATA
    if "eco" in aircraft_type.lower():
        return D328_ECO_DATA
    return DO328_100_DATA


# ----------------------------------------------------------------
# MOD VARIANT FACTORS (multiplier vs MOD 10 baseline)
# ----------------------------------------------------------------
# MOD 10: PW119B, reference configuration. All factors = 1.00
# MOD 20: PW119C (+5% thermodynamic power), higher single-engine ceiling.
#          Engine/propeller slightly higher costs, all else identical.
# MOD 30: PW119C + ground spoilers (standard) + 20deg flaps for T/O.
#          Better airfield performance. Adds spoiler system maintenance.
MOD_FACTORS = {
    "MOD 10": {
        "Airframe Checks": 1.00, "APU Inspections": 1.00, "FH-Based Tasks": 1.00,
        "Fatigue Damage (FD)": 1.00, "Corrosion Prevention (CP)": 1.00,
        "Structural Sampling (SSI)": 1.00, "Engines": 1.00, "Propellers": 1.00,
        "Landing Gear": 1.00, "APU": 1.00, "Time Controlled Components": 1.00,
    },
    "MOD 20": {
        "Airframe Checks": 1.00, "APU Inspections": 1.00, "FH-Based Tasks": 1.00,
        "Fatigue Damage (FD)": 1.00, "Corrosion Prevention (CP)": 1.00,
        "Structural Sampling (SSI)": 1.00, "Engines": 1.04, "Propellers": 1.02,
        "Landing Gear": 1.00, "APU": 1.00, "Time Controlled Components": 1.05,
    },
    "MOD 30": {
        "Airframe Checks": 1.03, "APU Inspections": 1.00, "FH-Based Tasks": 1.02,
        "Fatigue Damage (FD)": 1.00, "Corrosion Prevention (CP)": 1.00,
        "Structural Sampling (SSI)": 1.00, "Engines": 1.04, "Propellers": 1.02,
        "Landing Gear": 1.01, "APU": 1.00, "Time Controlled Components": 1.10,
    },
}
# Each environment/operational factor has different severity per category.
# Weight = how much of the base factor applies to this category.
# e.g., Tropical x1.12 with airframe weight 1.0 => full +12%
#        Tropical x1.12 with engine weight 0.3  => only +3.6%
# Formula per category: factor_applied = 1.0 + (base_factor - 1.0) * weight

# ENV_WEIGHTS[environment][category] = weight (0.0 to 1.5)
ENV_WEIGHTS = {
    "Temperate":        {"Airframe Checks": 1.0, "APU Inspections": 1.0, "FH-Based Tasks": 1.0, "Fatigue Damage (FD)": 1.0, "Corrosion Prevention (CP)": 1.0, "Structural Sampling (SSI)": 1.0, "Engines": 1.0, "Propellers": 1.0, "Landing Gear": 1.0, "APU": 1.0, "Time Controlled Components": 1.0},
    "Tropical / Humid": {"Airframe Checks": 1.3, "APU Inspections": 0.8, "FH-Based Tasks": 1.0, "Fatigue Damage (FD)": 0.6, "Corrosion Prevention (CP)": 1.5, "Structural Sampling (SSI)": 1.4, "Engines": 0.5, "Propellers": 0.4, "Landing Gear": 0.9, "APU": 0.7, "Time Controlled Components": 0.8},
    "Arid / Desert":    {"Airframe Checks": 0.5, "APU Inspections": 0.8, "FH-Based Tasks": 0.6, "Fatigue Damage (FD)": 0.3, "Corrosion Prevention (CP)": 0.2, "Structural Sampling (SSI)": 0.3, "Engines": 1.4, "Propellers": 1.5, "Landing Gear": 0.7, "APU": 0.9, "Time Controlled Components": 0.7},
    "Coastal / Marine":  {"Airframe Checks": 1.2, "APU Inspections": 0.7, "FH-Based Tasks": 0.9, "Fatigue Damage (FD)": 0.5, "Corrosion Prevention (CP)": 1.5, "Structural Sampling (SSI)": 1.5, "Engines": 0.4, "Propellers": 0.5, "Landing Gear": 1.3, "APU": 0.6, "Time Controlled Components": 0.7},
    "Cold / Arctic":     {"Airframe Checks": 0.6, "APU Inspections": 1.0, "FH-Based Tasks": 0.7, "Fatigue Damage (FD)": 0.4, "Corrosion Prevention (CP)": 0.3, "Structural Sampling (SSI)": 0.3, "Engines": 1.3, "Propellers": 0.8, "Landing Gear": 0.9, "APU": 1.2, "Time Controlled Components": 0.9},
    "High Altitude":     {"Airframe Checks": 0.2, "APU Inspections": 0.5, "FH-Based Tasks": 0.3, "Fatigue Damage (FD)": 0.3, "Corrosion Prevention (CP)": 0.1, "Structural Sampling (SSI)": 0.1, "Engines": 1.5, "Propellers": 1.3, "Landing Gear": 0.4, "APU": 0.6, "Time Controlled Components": 0.5},
}

# GRAVEL_WEIGHTS[category] = weight (how much gravel % affects this category)
GRAVEL_WEIGHTS = {
    "Airframe Checks": 0.6,
    "APU Inspections": 0.2,
    "FH-Based Tasks": 0.4,
    "Fatigue Damage (FD)": 0.3,
    "Corrosion Prevention (CP)": 0.5,
    "Structural Sampling (SSI)": 0.4,
    "Engines": 0.6,
    "Propellers": 1.5,
    "Landing Gear": 1.4,
    "APU": 0.2,
    "Time Controlled Components": 0.4,
}

# STOL_WEIGHTS[category] = weight (how much STOL % affects this category)
STOL_WEIGHTS = {
    "Airframe Checks": 0.5,
    "APU Inspections": 0.2,
    "FH-Based Tasks": 0.4,
    "Fatigue Damage (FD)": 1.5,
    "Corrosion Prevention (CP)": 0.2,
    "Structural Sampling (SSI)": 0.3,
    "Engines": 0.8,
    "Propellers": 1.2,
    "Landing Gear": 1.5,
    "APU": 0.3,
    "Time Controlled Components": 0.3,
}


def get_category_factor(category, env_mix, gravel_pct, stol_pct, mod_variant="MOD 10"):
    """Compute combined adjustment factor for a specific category using blended environment.
    The 5 operational environments (Temperate, Tropical, Arid, Coastal, Cold) are blended
    as a weighted average summing to 100%. High Altitude is applied as an independent
    percentage-based adder on top. MOD variant factor is applied multiplicatively."""
    # Blended environment factor from the 5 ops environments
    OPS_ENVS = ["Temperate", "Tropical / Humid", "Arid / Desert", "Coastal / Marine", "Cold / Arctic"]
    ops_total = sum(env_mix.get(e, 0) for e in OPS_ENVS)

    env_applied = 0.0
    if ops_total > 0:
        for env_name in OPS_ENVS:
            pct = env_mix.get(env_name, 0)
            if pct <= 0:
                continue
            weight_frac = pct / ops_total
            base_env = ENVIRONMENT_FACTORS.get(env_name, 1.0)
            env_weight = ENV_WEIGHTS.get(env_name, {}).get(category, 1.0)
            env_cat = 1.0 + (base_env - 1.0) * env_weight
            env_applied += env_cat * weight_frac
    else:
        env_applied = 1.0

    # High Altitude: independent adder based on percentage of ops at high altitude
    ha_pct = env_mix.get("High Altitude", 0)
    if ha_pct > 0:
        ha_base = ENVIRONMENT_FACTORS.get("High Altitude", 1.04)
        ha_weight = ENV_WEIGHTS.get("High Altitude", {}).get(category, 1.0)
        ha_adder = (ha_base - 1.0) * ha_weight * (ha_pct / 100)
        env_applied += ha_adder

    gravel_base = (gravel_pct / 100) * 0.15
    gravel_weight = GRAVEL_WEIGHTS.get(category, 1.0)
    gravel_applied = 1.0 + gravel_base * gravel_weight

    stol_base = (stol_pct / 100) * 0.10
    stol_weight = STOL_WEIGHTS.get(category, 1.0)
    stol_applied = 1.0 + stol_base * stol_weight

    # MOD variant factor
    mod_factor = MOD_FACTORS.get(mod_variant, {}).get(category, 1.0)

    combined = env_applied * gravel_applied * stol_applied * mod_factor
    return combined, env_applied, gravel_applied, stol_applied


# ----------------------------------------------------------------
# DMC CALCULATION ENGINE
# ----------------------------------------------------------------
def calculate_dmc(data, fh_yr, fc_yr, apu_hrs_yr, labour_rate, env_mix, gravel_pct, stol_pct, mod_variant="MOD 10"):
    results = []

    for item in data:
        int1 = item["int1"]
        param1 = item["param1"]
        int2 = item["int2"]
        param2 = item["param2"]
        mh = item["mh"]
        mat = item["mat"]
        cat = item["category"]

        # Category-specific combined factor (includes MOD variant)
        cat_factor, _, _, _ = get_category_factor(cat, env_mix, gravel_pct, stol_pct, mod_variant)

        # Occurrence 1 (calendar)
        occ1 = 0.0
        if int1 is not None and param1 is not None:
            p1 = param1.lower()
            if "day" in p1:
                occ1 = 365.0 / int1
            elif "month" in p1:
                occ1 = 12.0 / int1

        # Occurrence 2 (usage)
        occ2 = 0.0
        if int2 is not None and param2 is not None:
            p2 = param2.lower()
            if p2 == "fh":
                occ2 = fh_yr / int2
            elif p2 == "fc":
                occ2 = fc_yr / int2
            elif "apu" in p2:
                occ2 = apu_hrs_yr / int2
        elif int2 is not None and param2 is None:
            occ2 = fh_yr / int2

        occ = max(occ1, occ2)
        occ_source = "Calendar" if occ1 > occ2 else ("Usage" if occ2 > 0 else "Calendar")

        dmc_labour = (occ * labour_rate * mh) / fh_yr if fh_yr > 0 else 0
        dmc_material = (occ * mat) / fh_yr if fh_yr > 0 else 0

        dmc_labour_adj = dmc_labour * cat_factor
        dmc_material_adj = dmc_material * cat_factor
        dmc_total = dmc_labour_adj + dmc_material_adj

        results.append({
            "Category": cat,
            "Inspection": item["inspection"],
            "Interval 1": f"{int1} {param1}" if int1 and param1 else " -- ",
            "Interval 2": f"{int(int2)} {param2}" if int2 and param2 else " -- ",
            "MH": mh,
            "Material (EUR)": mat,
            "Occ/yr (Cal)": round(occ1, 4),
            "Occ/yr (Usage)": round(occ2, 4),
            "Occ/yr (Used)": round(occ, 4),
            "Driver": occ_source,
            "Adj. Factor": round(cat_factor, 4),
            "DMC Labour (EUR/FH)": round(dmc_labour_adj, 4),
            "DMC Material (EUR/FH)": round(dmc_material_adj, 4),
            "DMC Total (EUR/FH)": round(dmc_total, 4),
        })

    return results


# ----------------------------------------------------------------
# SESSION STATE
# ----------------------------------------------------------------
if "setup" not in st.session_state:
    st.session_state.setup = {
        "aircraft_type": "Do 328-100 (Turboprop)",
        "mod_variant": "MOD 10",
        "operator": "",
        "base_country": "",
        "fh_per_year": 2000,
        "fc_per_year": 2500,
        "fh_fc_ratio": 0.80,
        "apu_hrs_per_year": 2200,
        "env_mix": {"Temperate": 100, "Tropical / Humid": 0, "Arid / Desert": 0, "Coastal / Marine": 0, "Cold / Arctic": 0, "High Altitude": 0},
        "gravel_pct": 0,
        "labour_rate": 85.0,
        "stol_pct": 0,
    }

if "page" not in st.session_state:
    st.session_state.page = "Home"


# ----------------------------------------------------------------
# ICON SVGs (inline, no emoji)
# ----------------------------------------------------------------
def svg_icon(name, size=20):
    icons = {
        "plane": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M17.8 19.2L16 11l3.5-3.5C21 6 21.5 4 21 3c-1-.5-3 0-4.5 1.5L13 8 4.8 6.2c-.5-.1-.9.1-1.1.5l-.3.5c-.2.5-.1 1 .3 1.3L9 12l-2 3H4l-1 1 3 2 2 3 1-1v-3l3-2 3.5 5.3c.3.4.8.5 1.3.3l.5-.2c.4-.3.6-.7.5-1.2z"/></svg>',
        "settings": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 00.33 1.82l.06.06a2 2 0 010 2.83 2 2 0 01-2.83 0l-.06-.06a1.65 1.65 0 00-1.82-.33 1.65 1.65 0 00-1 1.51V21a2 2 0 01-2 2 2 2 0 01-2-2v-.09A1.65 1.65 0 009 19.4a1.65 1.65 0 00-1.82.33l-.06.06a2 2 0 01-2.83 0 2 2 0 010-2.83l.06-.06A1.65 1.65 0 004.68 15a1.65 1.65 0 00-1.51-1H3a2 2 0 01-2-2 2 2 0 012-2h.09A1.65 1.65 0 004.6 9a1.65 1.65 0 00-.33-1.82l-.06-.06a2 2 0 010-2.83 2 2 0 012.83 0l.06.06A1.65 1.65 0 009 4.68a1.65 1.65 0 001-1.51V3a2 2 0 012-2 2 2 0 012 2v.09a1.65 1.65 0 001 1.51 1.65 1.65 0 001.82-.33l.06-.06a2 2 0 012.83 0 2 2 0 010 2.83l-.06.06A1.65 1.65 0 0019.4 9a1.65 1.65 0 001.51 1H21a2 2 0 012 2 2 2 0 01-2 2h-.09a1.65 1.65 0 00-1.51 1z"/></svg>',
        "chart": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="18" y1="20" x2="18" y2="10"/><line x1="12" y1="20" x2="12" y2="4"/><line x1="6" y1="20" x2="6" y2="14"/></svg>',
        "file": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/></svg>',
        "wrench": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14.7 6.3a1 1 0 000 1.4l1.6 1.6a1 1 0 001.4 0l3.77-3.77a6 6 0 01-7.94 7.94l-6.91 6.91a2.12 2.12 0 01-3-3l6.91-6.91a6 6 0 017.94-7.94l-3.76 3.76z"/></svg>',
        "engine": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="3"/><path d="M12 1v2M12 21v2M4.22 4.22l1.42 1.42M18.36 18.36l1.42 1.42M1 12h2M21 12h2M4.22 19.78l1.42-1.42M18.36 5.64l1.42-1.42"/></svg>',
        "shield": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/></svg>',
        "globe": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><line x1="2" y1="12" x2="22" y2="12"/><path d="M12 2a15.3 15.3 0 014 10 15.3 15.3 0 01-4 10 15.3 15.3 0 01-4-10 15.3 15.3 0 014-10z"/></svg>',
        "dollar": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 000 7h5a3.5 3.5 0 010 7H6"/></svg>',
        "target": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><circle cx="12" cy="12" r="6"/><circle cx="12" cy="12" r="2"/></svg>',
        "layers": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polygon points="12 2 2 7 12 12 22 7 12 2"/><polyline points="2 17 12 22 22 17"/><polyline points="2 12 12 17 22 12"/></svg>',
        "zap": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"/></svg>',
        "search": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>',
        "download": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>',
        "check": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="#16a34a" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>',
        "cog": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 15a3 3 0 100-6 3 3 0 000 6z"/><path d="M19.4 15a1.65 1.65 0 00.33 1.82l.06.06a2 2 0 11-2.83 2.83l-.06-.06a1.65 1.65 0 00-1.82-.33 1.65 1.65 0 00-1 1.51V21a2 2 0 11-4 0v-.09A1.65 1.65 0 009 19.4a1.65 1.65 0 00-1.82.33l-.06.06a2 2 0 11-2.83-2.83l.06-.06A1.65 1.65 0 004.68 15a1.65 1.65 0 00-1.51-1H3a2 2 0 110-4h.09A1.65 1.65 0 004.6 9a1.65 1.65 0 00-.33-1.82l-.06-.06a2 2 0 112.83-2.83l.06.06A1.65 1.65 0 009 4.68a1.65 1.65 0 001-1.51V3a2 2 0 114 0v.09a1.65 1.65 0 001 1.51 1.65 1.65 0 001.82-.33l.06-.06a2 2 0 112.83 2.83l-.06.06A1.65 1.65 0 0019.4 9a1.65 1.65 0 001.51 1H21a2 2 0 110 4h-.09a1.65 1.65 0 00-1.51 1z"/></svg>',
        "wheel": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="10"/><circle cx="12" cy="12" r="4"/><line x1="12" y1="2" x2="12" y2="8"/><line x1="12" y1="16" x2="12" y2="22"/><line x1="2" y1="12" x2="8" y2="12"/><line x1="16" y1="12" x2="22" y2="12"/></svg>',
        "refresh": f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="23 4 23 10 17 10"/><polyline points="1 20 1 14 7 14"/><path d="M3.51 9a9 9 0 0114.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0020.49 15"/></svg>',
    }
    return icons.get(name, "")


# ----------------------------------------------------------------
# STYLING
# ----------------------------------------------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600&display=swap');

:root {
    --navy-900: #0A0F1E;
    --navy-800: #0F172A;
    --navy-700: #1E293B;
    --navy-600: #334155;
    --slate-500: #64748B;
    --slate-400: #94A3B8;
    --slate-300: #CBD5E1;
    --slate-200: #E2E8F0;
    --slate-100: #F1F5F9;
    --slate-50: #F8FAFC;
    --blue-600: #2563EB;
    --blue-500: #3B82F6;
    --blue-400: #60A5FA;
    --blue-100: #DBEAFE;
    --cyan-400: #22D3EE;
    --green-500: #22C55E;
    --green-100: #DCFCE7;
    --amber-500: #F59E0B;
    --amber-100: #FEF3C7;
}

.stApp {
    font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    background: var(--slate-50);
}

/* Sidebar */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, var(--navy-900) 0%, #0D1529 100%);
    border-right: 1px solid rgba(59,130,246,0.1);
}
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] .stMarkdown li,
section[data-testid="stSidebar"] .stMarkdown h1,
section[data-testid="stSidebar"] .stMarkdown h2,
section[data-testid="stSidebar"] .stMarkdown h3 {
    color: var(--slate-300) !important;
}
section[data-testid="stSidebar"] .stRadio label p {
    color: var(--slate-300) !important;
    font-weight: 500;
    font-size: 0.9rem;
}

/* Hero */
.hero {
    background: linear-gradient(135deg, var(--navy-900) 0%, #101D3A 40%, #152952 70%, var(--blue-600) 100%);
    padding: 2.75rem 3rem 2.25rem 3rem;
    border-radius: 16px;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
    border: 1px solid rgba(59,130,246,0.12);
    box-shadow: 0 4px 24px rgba(10,15,30,0.15);
}
.hero::before {
    content: '';
    position: absolute;
    top: -80px; right: -40px;
    width: 350px; height: 350px;
    background: radial-gradient(circle, rgba(59,130,246,0.06) 0%, transparent 70%);
}
.hero::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, var(--blue-400), var(--cyan-400), var(--blue-400), transparent);
}
.hero-brand {
    font-size: 0.7rem; font-weight: 700; letter-spacing: 4px;
    text-transform: uppercase; color: var(--blue-400); margin-bottom: 0.6rem;
    display: flex; align-items: center; gap: 6px;
}
.hero-title {
    font-size: 2.6rem; font-weight: 800; color: #FFF; line-height: 1.15; letter-spacing: -0.5px;
}
.hero-title span {
    background: linear-gradient(135deg, var(--blue-400), var(--cyan-400));
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.hero-sub {
    font-size: 1rem; color: var(--slate-400); font-weight: 400; margin-top: 0.6rem; line-height: 1.6;
    max-width: 650px;
}

/* Section */
.sec-head {
    font-size: 1.25rem; font-weight: 700; color: var(--navy-700);
    margin: 2.25rem 0 1rem 0; padding-bottom: 0.5rem;
    border-bottom: 2px solid var(--slate-200);
    display: flex; align-items: center; gap: 8px;
}
.sec-head span { color: var(--blue-600); }

/* Cards grid */
.card-grid {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(270px, 1fr));
    gap: 1rem; margin: 1.25rem 0;
}
.card {
    background: #FFF; border: 1px solid var(--slate-200); border-radius: 12px;
    padding: 1.5rem; transition: all 0.2s ease; position: relative; overflow: hidden;
}
.card:hover {
    border-color: var(--blue-500);
    box-shadow: 0 8px 30px rgba(37,99,235,0.07);
    transform: translateY(-1px);
}
.card-accent {
    position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, var(--blue-600), var(--cyan-400));
    opacity: 0; transition: opacity 0.2s;
}
.card:hover .card-accent { opacity: 1; }
.card-icon {
    width: 40px; height: 40px; border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    margin-bottom: 0.85rem; color: var(--blue-600);
}
.card-icon.blue { background: var(--blue-100); color: var(--blue-600); }
.card-icon.green { background: var(--green-100); color: #16a34a; }
.card-icon.amber { background: var(--amber-100); color: #d97706; }
.card-icon.slate { background: var(--slate-100); color: var(--navy-600); }
.card-t { font-size: 0.95rem; font-weight: 700; color: var(--navy-700); margin-bottom: 0.3rem; }
.card-d { font-size: 0.82rem; color: var(--slate-500); line-height: 1.55; }

/* Metric boxes */
.metrics { display: flex; gap: 0.85rem; margin: 1.5rem 0; flex-wrap: wrap; }
.metric {
    flex: 1; min-width: 170px;
    background: linear-gradient(135deg, var(--navy-900), #111D35);
    border-radius: 12px; padding: 1.15rem; text-align: center;
    border: 1px solid rgba(59,130,246,0.15);
    box-shadow: 0 2px 12px rgba(10,15,30,0.2);
}
.metric-label {
    font-size: 0.62rem; font-weight: 700; letter-spacing: 2px;
    text-transform: uppercase; color: var(--blue-400); margin-bottom: 0.2rem;
}
.metric-val {
    font-family: 'JetBrains Mono', monospace; font-size: 1.35rem;
    font-weight: 700; color: #FFF;
}
.metric-unit { font-size: 0.7rem; color: var(--slate-400); }

/* Setup card */
.s-card {
    background: #FFF; border: 1px solid var(--slate-200); border-radius: 12px;
    padding: 1.5rem; margin-bottom: 1rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.03);
}
.s-card-t {
    display: flex; align-items: center; gap: 8px;
    font-size: 0.95rem; font-weight: 700; color: var(--navy-800);
    margin-bottom: 1rem; padding-bottom: 0.65rem;
    border-bottom: 1px solid var(--slate-100);
}

/* Chips */
.chips { display: flex; flex-wrap: wrap; gap: 0.4rem; margin: 0.75rem 0; }
.chip {
    background: var(--blue-100); border: 1px solid #93C5FD; border-radius: 8px;
    padding: 0.3rem 0.7rem; font-size: 0.72rem; font-weight: 600; color: #1E40AF;
}

/* Info boxes */
.info-box {
    border-radius: 8px; padding: 0.65rem 1rem; font-size: 0.78rem; margin: 0.5rem 0;
}
.info-blue { background: #EFF6FF; border: 1px solid #BFDBFE; color: #1E40AF; }
.info-green { background: #F0FDF4; border: 1px solid #86EFAC; color: #166534; }
.info-amber { background: #FFFBEB; border: 1px solid #FDE68A; color: #92400E; }
.info-slate { background: var(--slate-50); border: 1px solid var(--slate-200); color: var(--slate-500); }

/* Footer */
.footer {
    text-align: center; padding: 1.75rem 0 1rem 0; margin-top: 3rem;
    border-top: 1px solid var(--slate-200); color: var(--slate-400); font-size: 0.75rem;
}
.footer strong { color: var(--slate-500); }

/* Step number */
.step-num {
    width: 32px; height: 32px; border-radius: 8px;
    background: var(--blue-600); color: #FFF;
    display: flex; align-items: center; justify-content: center;
    font-weight: 700; font-size: 0.85rem; margin-bottom: 0.75rem;
}
</style>
""", unsafe_allow_html=True)


# ----------------------------------------------------------------
# DATA
# ----------------------------------------------------------------
COUNTRIES = [
    "Germany", "France", "United Kingdom", "United States", "Canada",
    "Australia", "Brazil", "Guinea", "Nigeria", "South Africa",
    "India", "Japan", "South Korea", "China", "UAE",
    "Saudi Arabia", "Turkey", "Italy", "Spain", "Sweden",
    "Norway", "Switzerland", "Netherlands", "Austria", "Poland",
    "Czech Republic", "Greece", "Portugal", "Mexico", "Argentina",
    "Chile", "Colombia", "Peru", "Egypt", "Kenya",
    "Tanzania", "Ethiopia", "Morocco", "Thailand", "Indonesia",
    "Malaysia", "Philippines", "Vietnam", "New Zealand", "Singapore", "Pakistan", "Other"
]

ENVIRONMENTS = ["Temperate", "Tropical / Humid", "Arid / Desert", "Coastal / Marine", "Cold / Arctic", "High Altitude"]

AIRCRAFT_TYPES = {
    "D328eco": {
        "full_name": "D328eco (Next-Gen Turboprop)", "engines": "PW127S-XT (x2)",
        "pax": "40", "description": "Next-generation regional turboprop with modern avionics.",
    },
    "Do 328-100 (Turboprop)": {
        "full_name": "Dornier 328-100 Turboprop", "engines": "PW119B / PW119C (x2)",
        "pax": "32", "description": "Proven regional turboprop with PW100-series engines.",
    },
    "Do 328-300 (Jet)": {
        "full_name": "Dornier 328-300 JET", "engines": "PW306B (x2)",
        "pax": "32-34", "description": "Jet variant with rear-mounted PW306B turbofans.",
    },
}

ENVIRONMENT_FACTORS = {
    "Temperate": 1.00, "Tropical / Humid": 1.12, "Arid / Desert": 1.08,
    "Coastal / Marine": 1.10, "Cold / Arctic": 1.06, "High Altitude": 1.04,
}


# ----------------------------------------------------------------
# SIDEBAR
# ----------------------------------------------------------------
with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 1.75rem 0 1.25rem 0;">
        <div style="width:48px; height:48px; background:linear-gradient(135deg, #2563EB, #22D3EE); border-radius:12px; display:flex; align-items:center; justify-content:center; margin:0 auto 0.6rem auto; box-shadow: 0 4px 12px rgba(37,99,235,0.3);">
            <svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><path d="M17.8 19.2L16 11l3.5-3.5C21 6 21.5 4 21 3c-1-.5-3 0-4.5 1.5L13 8 4.8 6.2c-.5-.1-.9.1-1.1.5l-.3.5c-.2.5-.1 1 .3 1.3L9 12l-2 3H4l-1 1 3 2 2 3 1-1v-3l3-2 3.5 5.3c.3.4.8.5 1.3.3l.5-.2c.4-.3.6-.7.5-1.2z"/></svg>
        </div>
        <div style="font-size:1.35rem; font-weight:800; color:#FFFFFF; letter-spacing:-0.5px;">
            Maint<span style="background:linear-gradient(135deg,#60A5FA,#22D3EE);-webkit-background-clip:text;-webkit-text-fill-color:transparent;">Edge</span>
        </div>
        <div style="font-size:0.6rem; font-weight:600; letter-spacing:2.5px; color:#475569; text-transform:uppercase; margin-top:3px;">
            DMC Calculator
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    page = st.radio("Navigation", ["Home", "Setup & Calculate", "Report"], label_visibility="collapsed")
    st.session_state.page = page
    st.markdown("---")
    st.markdown("""
    <div style="text-align:center; padding:0.75rem 0;">
        <div style="font-size:0.55rem; letter-spacing:2.5px; color:#475569; text-transform:uppercase; margin-bottom:4px;">Powered by</div>
        <div style="font-size:0.8rem; font-weight:700; color:#94A3B8;">Deutsche Aircraft GmbH</div>
        <div style="font-size:0.6rem; color:#475569; margin-top:3px;">v1.0.0</div>
    </div>
    """, unsafe_allow_html=True)


# ================================================================
# HOME
# ================================================================
if st.session_state.page == "Home":

    st.markdown(f"""
    <div class="hero">
        <div class="hero-brand">{svg_icon("plane", 14)} Deutsche Aircraft GmbH</div>
        <div class="hero-title">Maint<span>Edge</span></div>
        <div class="hero-title" style="font-size:1.45rem; font-weight:400; color:#CBD5E1; margin-top:0.15rem; letter-spacing:0;">
            Direct Maintenance Cost Calculator
        </div>
        <div class="hero-sub">
            Precision DMC modeling for the D328 family. From line maintenance to heavy checks,
            engine shop visits to component overhauls -- fully parameterized for your operation.
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="metrics">
        <div class="metric"><div class="metric-label">Aircraft Variants</div><div class="metric-val">3</div><div class="metric-unit">D328eco / 328-100 / 328JET</div></div>
        <div class="metric"><div class="metric-label">Maintenance Items</div><div class="metric-val">58+</div><div class="metric-unit">Checks / FD / CP / SSI / Heavy</div></div>
        <div class="metric"><div class="metric-label">Output</div><div class="metric-val">EUR/FH</div><div class="metric-unit">Labour + Material split</div></div>
        <div class="metric"><div class="metric-label">Adjustments</div><div class="metric-val">6</div><div class="metric-unit">Environment & ops factors</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f'<div class="sec-head">{svg_icon("wrench", 20)} <span>Core</span> Capabilities</div>', unsafe_allow_html=True)

    st.markdown(f"""
    <div class="card-grid">
        <div class="card"><div class="card-accent"></div>
            <div class="card-icon blue">{svg_icon("plane", 22)}</div>
            <div class="card-t">Airframe Checks</div>
            <div class="card-d">LC, A-checks (A1-A5), C-checks (C1-C4) with dual calendar/FH intervals. Automatic occurrence selection using the higher-frequency driver.</div>
        </div>
        <div class="card"><div class="card-accent"></div>
            <div class="card-icon amber">{svg_icon("engine", 22)}</div>
            <div class="card-t">Engine & Propeller</div>
            <div class="card-d">Engine shop visit at 8,000 FH and propeller change at 6,000 FC. Full labour + material amortization.</div>
        </div>
        <div class="card"><div class="card-accent"></div>
            <div class="card-icon green">{svg_icon("shield", 22)}</div>
            <div class="card-t">Structural & Fatigue</div>
            <div class="card-d">17 Fatigue Damage (FD), 6 Corrosion Prevention (CP), and 5 Structural Sampling (SSI) inspections.</div>
        </div>
        <div class="card"><div class="card-accent"></div>
            <div class="card-icon slate">{svg_icon("wheel", 22)}</div>
            <div class="card-t">Landing Gear & APU</div>
            <div class="card-d">LG overhaul (144mo/22,000FC), brakes (3,000FC), APU overhaul (8,000FC), plus 4 APU inspection tasks.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f'<div class="sec-head">{svg_icon("zap", 20)} <span>Getting</span> Started</div>', unsafe_allow_html=True)

    st.markdown(f"""
    <div class="card-grid">
        <div class="card"><div class="card-accent"></div>
            <div class="step-num">1</div>
            <div class="card-t">Setup</div>
            <div class="card-d">Select aircraft type, operator, utilization (FH/FC/APU hrs), environment, gravel %, labour rate, and STOL %.</div>
        </div>
        <div class="card"><div class="card-accent"></div>
            <div class="step-num">2</div>
            <div class="card-t">Calculate</div>
            <div class="card-d">Engine computes occurrences, selects the limiting interval, and calculates EUR/FH for every maintenance item.</div>
        </div>
        <div class="card"><div class="card-accent"></div>
            <div class="step-num">3</div>
            <div class="card-t">Report</div>
            <div class="card-d">View full breakdown by category, export detailed tables, and download CSV reports.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="chips">
        <span class="chip">D328eco -- Next-Gen Turboprop</span>
        <span class="chip">Do 328-100 -- Turboprop (PW119)</span>
        <span class="chip">Do 328-300 -- JET (PW306B)</span>
    </div>
    """, unsafe_allow_html=True)



# ================================================================
# SETUP & CALCULATE (merged)
# ================================================================
elif st.session_state.page == "Setup & Calculate":

    st.markdown(f"""
    <div class="hero" style="padding:2rem 2.5rem 1.5rem 2.5rem;">
        <div class="hero-brand">{svg_icon("settings", 14)} Configuration & Analysis</div>
        <div class="hero-title" style="font-size:2rem;">Setup & <span>Calculate</span></div>
        <div class="hero-sub">Define your operational parameters, then calculate DMC with one click.</div>
    </div>
    """, unsafe_allow_html=True)

    s = st.session_state.setup
    col1, col2 = st.columns(2)

    with col1:
        st.markdown(f'<div class="s-card"><div class="s-card-t">{svg_icon("plane", 18)} Aircraft & Operator</div></div>', unsafe_allow_html=True)
        aircraft_type = st.selectbox("Aircraft Type", list(AIRCRAFT_TYPES.keys()),
            index=list(AIRCRAFT_TYPES.keys()).index(s["aircraft_type"]))
        s["aircraft_type"] = aircraft_type
        ac = AIRCRAFT_TYPES[aircraft_type]

        # MOD variant for Turboprop
        if "Turboprop" in aircraft_type or "328-100" in aircraft_type:
            mod_opts = ["MOD 10", "MOD 20", "MOD 30"]
            mod_idx = mod_opts.index(s.get("mod_variant", "MOD 10")) if s.get("mod_variant", "MOD 10") in mod_opts else 0
            s["mod_variant"] = st.selectbox("Configuration Variant", mod_opts, index=mod_idx,
                help="MOD 10: Base config | MOD 20: Enhanced avionics | MOD 30: Full modernisation")
            st.markdown(f'<div class="info-box info-slate"><strong>{ac["full_name"]} -- {s["mod_variant"]}</strong><br>Engines: {ac["engines"]} | Pax: {ac["pax"]}<br><span style="color:#94A3B8">{ac["description"]}</span></div>', unsafe_allow_html=True)
        else:
            s["mod_variant"] = "N/A"
            st.markdown(f'<div class="info-box info-slate"><strong>{ac["full_name"]}</strong><br>Engines: {ac["engines"]} | Pax: {ac["pax"]}<br><span style="color:#94A3B8">{ac["description"]}</span></div>', unsafe_allow_html=True)

        s["operator"] = st.text_input("Operator Name", value=s["operator"], placeholder="e.g. UMSI Guinea, Nolinor Aviation")
        s["base_country"] = st.selectbox("Base Country", [""] + COUNTRIES,
            index=(COUNTRIES.index(s["base_country"]) + 1) if s["base_country"] in COUNTRIES else 0)

    with col2:
        st.markdown(f'<div class="s-card"><div class="s-card-t">{svg_icon("chart", 18)} Utilization Profile</div></div>', unsafe_allow_html=True)
        s["fh_per_year"] = st.number_input("Flight Hours / Year (FH/yr)", min_value=100, max_value=5000, value=s["fh_per_year"], step=50)
        s["fc_per_year"] = st.number_input("Flight Cycles / Year (FC/yr)", min_value=100, max_value=6000, value=s["fc_per_year"], step=50)
        if "eco" not in s["aircraft_type"].lower():
            s["apu_hrs_per_year"] = st.number_input("APU Hours / Year", min_value=100, max_value=5000, value=s["apu_hrs_per_year"], step=50,
                help="Annual APU operating hours for APU inspection intervals.")
        else:
            s["apu_hrs_per_year"] = 0
            st.markdown('<div class="info-box info-slate">No APU fitted -- D328eco does not have an APU.</div>', unsafe_allow_html=True)

        if s["fc_per_year"] > 0:
            ratio = round(s["fh_per_year"] / s["fc_per_year"], 2)
            avg_flt = round(s["fh_per_year"] / s["fc_per_year"] * 60, 0)
        else:
            ratio = 0.0
            avg_flt = 0
        s["fh_fc_ratio"] = ratio
        profile = "Short-haul" if ratio < 1.0 else ("Medium-haul" if ratio < 1.5 else "Long-haul")
        st.markdown(f'<div class="info-box info-blue"><strong>FH/FC Ratio: {ratio}</strong> | {profile} | Avg: {avg_flt:.0f} min/flight</div>', unsafe_allow_html=True)

    st.markdown("---")
    col3, col4 = st.columns(2)

    with col3:
        st.markdown(f'<div class="s-card"><div class="s-card-t">{svg_icon("globe", 18)} Operating Environment Mix</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="info-box info-blue">Set the percentage of operations per climate zone. Total must equal 100%.</div>', unsafe_allow_html=True)

        env_mix = s.get("env_mix", {"Temperate": 100, "Tropical / Humid": 0, "Arid / Desert": 0, "Coastal / Marine": 0, "Cold / Arctic": 0, "High Altitude": 0})
        # Only these 5 are the operational environment mix (must sum to 100%)
        OPS_ENVS = ["Temperate", "Tropical / Humid", "Arid / Desert", "Coastal / Marine", "Cold / Arctic"]
        ops_labels = {"Temperate": "Temperate (baseline)", "Tropical / Humid": "Tropical / Humid (+12%)",
            "Arid / Desert": "Arid / Desert (+8%)", "Coastal / Marine": "Coastal / Marine (+10%)",
            "Cold / Arctic": "Cold / Arctic (+6%)"}

        new_mix = {}
        for env_key in OPS_ENVS:
            new_mix[env_key] = st.slider(ops_labels[env_key], 0, 100, env_mix.get(env_key, 0), 5, key=f"env_{env_key}")

        ops_total = sum(new_mix.values())

        if ops_total == 100:
            blended_base = sum((ENVIRONMENT_FACTORS[e] * p / 100) for e, p in new_mix.items() if p > 0)
            active_envs = [f"{e.split('/')[0].strip()} {p}%" for e, p in new_mix.items() if p > 0]
            st.markdown(f'<div class="info-box info-green">Blended Env Factor: <strong>x{blended_base:.3f}</strong> -- {" | ".join(active_envs)}</div>', unsafe_allow_html=True)
        elif ops_total == 0:
            st.markdown(f'<div class="info-box info-amber">Set at least one environment percentage.</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="info-box info-amber">Total is <strong>{ops_total}%</strong> -- must equal 100%. Adjust sliders.</div>', unsafe_allow_html=True)

        # Keep High Altitude in the mix dict (set separately below) but not in the 100% constraint
        new_mix["High Altitude"] = env_mix.get("High Altitude", 0)
        s["env_mix"] = new_mix

    with col4:
        st.markdown(f'<div class="s-card"><div class="s-card-t">{svg_icon("dollar", 18)} Cost & Operations</div></div>', unsafe_allow_html=True)
        s["labour_rate"] = st.number_input("Labour Cost / Hour (EUR/hr)", min_value=20.0, max_value=250.0,
            value=s["labour_rate"], step=5.0, format="%.2f")
        if s["labour_rate"] < 50:
            lr = "Low-cost region"
        elif s["labour_rate"] < 100:
            lr = "Mid-range (typical EASA/FAA MRO)"
        else:
            lr = "Premium (Western Europe / North America)"
        st.markdown(f'<div class="info-box info-slate">{lr}</div>', unsafe_allow_html=True)

        st.markdown(f'<div class="s-card"><div class="s-card-t">{svg_icon("target", 18)} Operational Conditions</div></div>', unsafe_allow_html=True)
        s["stol_pct"] = st.slider("STOL Operations (%)", 0, 100, s["stol_pct"], 5)
        ha_pct = st.slider("High Altitude Operations (%)", 0, 100, env_mix.get("High Altitude", 0), 5,
            help="Percentage of operations from airfields above 5,000 ft. Applied as independent adder (+4% base).")
        s["env_mix"]["High Altitude"] = ha_pct
        s["gravel_pct"] = st.slider("Gravel Runway Operations (%)", 0, 100, s["gravel_pct"], 5)

    st.session_state.setup = s

    # Compute factors
    env_mix = s.get("env_mix", {"Temperate": 100})
    gf = 1.0 + (s["gravel_pct"] / 100) * 0.15
    sf = 1.0 + (s["stol_pct"] / 100) * 0.10
    blended_env = sum((ENVIRONMENT_FACTORS[e] * p / 100) for e, p in env_mix.items() if p > 0) if sum(env_mix.values()) == 100 else 1.0
    active_envs_short = ", ".join([f"{e.split('/')[0].strip()} {p}%" for e, p in env_mix.items() if p > 0])
    mod_label = s.get("mod_variant", "N/A")
    mod_display = f'<div class="metric-unit">{mod_label}</div>' if mod_label != "N/A" else ""

    # Summary bar
    st.markdown("---")
    st.markdown(f"""
    <div class="metrics">
        <div class="metric"><div class="metric-label">Aircraft</div><div class="metric-val" style="font-size:0.95rem;">{s["aircraft_type"]}</div>{mod_display}<div class="metric-unit">{s["operator"] or "N/A"} | {s["base_country"] or "N/A"}</div></div>
        <div class="metric"><div class="metric-label">Utilization</div><div class="metric-val" style="font-size:1rem;">{s["fh_per_year"]:,} FH / {s["fc_per_year"]:,} FC</div><div class="metric-unit">{"No APU" if "eco" in s["aircraft_type"].lower() else f"APU: {s['apu_hrs_per_year']:,} hrs"} | Ratio: {ratio}</div></div>
        <div class="metric"><div class="metric-label">Blended Env</div><div class="metric-val">x{blended_env:.3f}</div><div class="metric-unit">{active_envs_short}</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── CALCULATE BUTTON ──
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    with col_btn2:
        calculate_pressed = st.button(
            "CALCULATE DMC",
            use_container_width=True,
            type="primary",
        )

    if calculate_pressed:
        st.session_state.calculated = True

    if st.session_state.get("calculated", False):
        with st.spinner("Computing DMC for all maintenance items..."):
            results = calculate_dmc(get_aircraft_data(s["aircraft_type"]), s["fh_per_year"], s["fc_per_year"],
                s["apu_hrs_per_year"], s["labour_rate"], s.get("env_mix", {"Temperate": 100}), s["gravel_pct"], s["stol_pct"], s.get("mod_variant", "MOD 10"))

            # D328eco: inject PWC FMP engine DMC as a fixed EUR/FH item
            # The FMP rate already covers all engine maintenance; no environmental adjustment applied.
            if "eco" in s["aircraft_type"].lower():
                avg_min_disp = round((s["fh_per_year"] / s["fc_per_year"]) * 60, 1) if s["fc_per_year"] > 0 else 0
                pwc_rate = get_pwc_engine_rate_eur(s["fh_per_year"], s["fc_per_year"])
                results.append({
                    "Category":               "Engines (PWC FMP)",
                    "Inspection":             f"PW127XT-S FMP -- 2EA (avg {avg_min_disp} min/flight, $2023 escalated)",
                    "Interval 1":             "Pay-per-hour",
                    "Interval 2":             "Pay-per-hour",
                    "MH":                     0,
                    "Material (EUR)":         0,
                    "Occ/yr (Cal)":           0,
                    "Occ/yr (Usage)":         0,
                    "Occ/yr (Used)":          0,
                    "Driver":                 "FMP Rate",
                    "Adj. Factor":            1.0,
                    "DMC Labour (EUR/FH)":    0.0,
                    "DMC Material (EUR/FH)":  round(pwc_rate, 4),
                    "DMC Total (EUR/FH)":     round(pwc_rate, 4),
                })

            df = pd.DataFrame(results)
            st.session_state.calc_results = results

        total_labour = df["DMC Labour (EUR/FH)"].sum()
        total_material = df["DMC Material (EUR/FH)"].sum()
        total_dmc = total_labour + total_material

        st.markdown("---")

        # Result metrics with highlighted total
        st.markdown(f"""
        <div class="metrics">
            <div class="metric" style="border: 1px solid rgba(34,197,94,0.3); box-shadow: 0 4px 20px rgba(34,197,94,0.15);">
                <div class="metric-label" style="color:#4ade80;">Total DMC</div>
                <div class="metric-val" style="font-size:1.6rem;">EUR {total_dmc:,.2f}</div>
                <div class="metric-unit">per Flight Hour</div>
            </div>
            <div class="metric">
                <div class="metric-label">Labour DMC</div>
                <div class="metric-val">EUR {total_labour:,.2f}</div>
                <div class="metric-unit">per Flight Hour</div>
            </div>
            <div class="metric">
                <div class="metric-label">Material DMC</div>
                <div class="metric-val">EUR {total_material:,.2f}</div>
                <div class="metric-unit">per Flight Hour</div>
            </div>
            <div class="metric">
                <div class="metric-label">Annual DMC</div>
                <div class="metric-val">EUR {total_dmc * s["fh_per_year"]:,.0f}</div>
                <div class="metric-unit">at {s["fh_per_year"]:,} FH/yr</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Labour vs Material split bar
        labour_pct = (total_labour / total_dmc * 100) if total_dmc > 0 else 0
        material_pct = 100 - labour_pct
        st.markdown(f"""
        <div style="margin:1.5rem 0;">
            <div style="display:flex; justify-content:space-between; font-size:0.75rem; font-weight:600; color:#475569; margin-bottom:4px;">
                <span>Labour: {labour_pct:.1f}%</span>
                <span>Material: {material_pct:.1f}%</span>
            </div>
            <div style="height:10px; border-radius:5px; background:#E2E8F0; overflow:hidden; display:flex;">
                <div style="width:{labour_pct}%; background:linear-gradient(90deg, #2563EB, #3B82F6); border-radius:5px 0 0 5px;"></div>
                <div style="width:{material_pct}%; background:linear-gradient(90deg, #22D3EE, #06B6D4); border-radius:0 5px 5px 0;"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Category breakdown
        st.markdown(f'<div class="sec-head">{svg_icon("layers", 20)} <span>Category</span> Breakdown</div>', unsafe_allow_html=True)

        cat_sum = df.groupby("Category").agg({
            "DMC Labour (EUR/FH)": "sum", "DMC Material (EUR/FH)": "sum", "DMC Total (EUR/FH)": "sum"
        }).reset_index()
        cat_sum = cat_sum.sort_values("DMC Total (EUR/FH)", ascending=False)

        col_c1, col_c2 = st.columns(2)
        with col_c1:
            fig_pie = go.Figure(data=[go.Pie(
                labels=cat_sum["Category"], values=cat_sum["DMC Total (EUR/FH)"].round(2),
                hole=0.5, marker=dict(colors=["#2563EB", "#0891B2", "#059669", "#D97706", "#7C3AED", "#DC2626", "#475569", "#EC4899", "#F59E0B"]),
                textinfo="label+percent", textfont_size=10, textposition="outside",
            )])
            fig_pie.update_layout(title=dict(text="DMC Distribution by Category", font=dict(size=13, family="Plus Jakarta Sans")),
                height=420, margin=dict(t=50, b=20, l=20, r=20), showlegend=False,
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_pie, use_container_width=True)

        with col_c2:
            fig_bar = go.Figure()
            fig_bar.add_trace(go.Bar(name="Labour", x=cat_sum["Category"], y=cat_sum["DMC Labour (EUR/FH)"].round(2), marker_color="#2563EB"))
            fig_bar.add_trace(go.Bar(name="Material", x=cat_sum["Category"], y=cat_sum["DMC Material (EUR/FH)"].round(2), marker_color="#22D3EE"))
            fig_bar.update_layout(title=dict(text="Labour vs Material (EUR/FH)", font=dict(size=13, family="Plus Jakarta Sans")),
                barmode="stack", height=420, margin=dict(t=50, b=100, l=50, r=20),
                xaxis=dict(tickangle=-40, tickfont=dict(size=9)),
                yaxis=dict(title=dict(text="EUR/FH", font=dict(size=11))),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_bar, use_container_width=True)

        # Category table
        st.markdown(f'<div class="sec-head">{svg_icon("file", 20)} <span>Category</span> Summary</div>', unsafe_allow_html=True)
        cat_disp = cat_sum.copy()
        cat_disp["% of Total"] = (cat_disp["DMC Total (EUR/FH)"] / total_dmc * 100).round(1)

        # Add per-category adjustment factors
        cat_factors_list = []
        for cat_name in cat_disp["Category"]:
            cf, _, _, _ = get_category_factor(cat_name, s.get("env_mix", {"Temperate": 100}), s["gravel_pct"], s["stol_pct"], s.get("mod_variant", "MOD 10"))
            cat_factors_list.append(round(cf, 4))
        cat_disp["Adj. Factor"] = cat_factors_list

        st.dataframe(cat_disp.style.format({
            "DMC Labour (EUR/FH)": "EUR {:.2f}", "DMC Material (EUR/FH)": "EUR {:.2f}",
            "DMC Total (EUR/FH)": "EUR {:.2f}", "% of Total": "{:.1f}%", "Adj. Factor": "x{:.4f}"}),
            use_container_width=True, hide_index=True)

        # Category factor breakdown
        st.markdown(f'<div class="sec-head">{svg_icon("target", 20)} <span>Category-Specific</span> Adjustment Factors</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="info-box info-blue">Factors are weighted per category. For example, tropical humidity affects Corrosion Prevention tasks more heavily than engine tasks.</div>', unsafe_allow_html=True)

        factor_rows = []
        all_cats = sorted(df["Category"].unique().tolist())
        for cat_name in all_cats:
            cf, ef_cat, gf_cat, sf_cat = get_category_factor(cat_name, s.get("env_mix", {"Temperate": 100}), s["gravel_pct"], s["stol_pct"], s.get("mod_variant", "MOD 10"))
            factor_rows.append({
                "Category": cat_name,
                "Env Factor": round(ef_cat, 4),
                "Gravel Factor": round(gf_cat, 4),
                "STOL Factor": round(sf_cat, 4),
                "Combined": round(cf, 4),
            })
        factor_df = pd.DataFrame(factor_rows)
        st.dataframe(factor_df.style.format({
            "Env Factor": "x{:.4f}", "Gravel Factor": "x{:.4f}",
            "STOL Factor": "x{:.4f}", "Combined": "x{:.4f}"}),
            use_container_width=True, hide_index=True)

        # Detail table
        st.markdown(f'<div class="sec-head">{svg_icon("search", 20)} <span>Detailed</span> Item Breakdown</div>', unsafe_allow_html=True)
        categories = ["All"] + sorted(df["Category"].unique().tolist())
        sel_cat = st.selectbox("Filter by Category", categories)
        df_show = df if sel_cat == "All" else df[df["Category"] == sel_cat]

        st.dataframe(df_show.style.format({
            "Material (EUR)": "EUR {:,.2f}", "Occ/yr (Cal)": "{:.4f}", "Occ/yr (Usage)": "{:.4f}",
            "Occ/yr (Used)": "{:.4f}", "Adj. Factor": "{:.4f}", "DMC Labour (EUR/FH)": "EUR {:.4f}",
            "DMC Material (EUR/FH)": "EUR {:.4f}", "DMC Total (EUR/FH)": "EUR {:.4f}"}),
            use_container_width=True, hide_index=True, height=600)


    else:
        # Pre-calculation state
        st.markdown("""
        <div style="text-align:center; padding:3rem 0; color:#94A3B8;">
            <div style="font-size:3rem; margin-bottom:0.5rem;">
                <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="#CBD5E1" stroke-width="1.5"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg>
            </div>
            <div style="font-size:1.1rem; font-weight:600; color:#64748B;">Ready to calculate</div>
            <div style="font-size:0.85rem; margin-top:0.25rem;">Verify your parameters above, then press CALCULATE DMC</div>
        </div>
        """, unsafe_allow_html=True)


# ================================================================
# REPORT
# ================================================================
elif st.session_state.page == "Report":

    st.markdown(f"""
    <div class="hero" style="padding:2rem 2.5rem 1.5rem 2.5rem;">
        <div class="hero-brand">{svg_icon("download", 14)} Export</div>
        <div class="hero-title" style="font-size:2rem;">DMC <span>Report</span></div>
        <div class="hero-sub">Generate and download professional DMC reports in PDF and Excel formats.</div>
    </div>
    """, unsafe_allow_html=True)

    s = st.session_state.setup
    if s["fh_per_year"] == 0 or s["fc_per_year"] == 0:
        st.error("Please complete Setup and Calculate first.")
        st.stop()

    env_mix = s.get("env_mix", {"Temperate": 100})
    gf = 1.0 + (s["gravel_pct"] / 100) * 0.15
    sf = 1.0 + (s["stol_pct"] / 100) * 0.10
    blended_env = sum((ENVIRONMENT_FACTORS[e] * p / 100) for e, p in env_mix.items() if p > 0) if sum(env_mix.values()) == 100 else 1.0
    active_envs_str = ", ".join([f"{e} {p}%" for e, p in env_mix.items() if p > 0])

    results = calculate_dmc(get_aircraft_data(s["aircraft_type"]), s["fh_per_year"], s["fc_per_year"],
        s["apu_hrs_per_year"], s["labour_rate"], env_mix, s["gravel_pct"], s["stol_pct"], s.get("mod_variant", "MOD 10"))

    if "eco" in s["aircraft_type"].lower():
        avg_min_disp = round((s["fh_per_year"] / s["fc_per_year"]) * 60, 1) if s["fc_per_year"] > 0 else 0
        pwc_rate = get_pwc_engine_rate_eur(s["fh_per_year"], s["fc_per_year"])
        results.append({
            "Category":               "Engines (PWC FMP)",
            "Inspection":             f"PW127XT-S FMP -- 2EA (avg {avg_min_disp} min/flight, $2023 escalated)",
            "Interval 1":             "Pay-per-hour",
            "Interval 2":             "Pay-per-hour",
            "MH":                     0,
            "Material (EUR)":         0,
            "Occ/yr (Cal)":           0,
            "Occ/yr (Usage)":         0,
            "Occ/yr (Used)":          0,
            "Driver":                 "FMP Rate",
            "Adj. Factor":            1.0,
            "DMC Labour (EUR/FH)":    0.0,
            "DMC Material (EUR/FH)":  round(pwc_rate, 4),
            "DMC Total (EUR/FH)":     round(pwc_rate, 4),
        })

    df = pd.DataFrame(results)

    total_labour = df["DMC Labour (EUR/FH)"].sum()
    total_material = df["DMC Material (EUR/FH)"].sum()
    total_dmc = total_labour + total_material

    st.markdown(f"""
    <div class="metrics">
        <div class="metric"><div class="metric-label">Total DMC</div><div class="metric-val">EUR {total_dmc:,.2f} /FH</div></div>
        <div class="metric"><div class="metric-label">Labour</div><div class="metric-val">EUR {total_labour:,.2f} /FH</div></div>
        <div class="metric"><div class="metric-label">Material</div><div class="metric-val">EUR {total_material:,.2f} /FH</div></div>
        <div class="metric"><div class="metric-label">Annual</div><div class="metric-val">EUR {total_dmc * s['fh_per_year']:,.0f}</div></div>
    </div>
    """, unsafe_allow_html=True)

    op_name = s["operator"].replace(" ", "_") if s["operator"] else "Generic"
    base_filename = f"MaintEdge_DMC_{s['aircraft_type'].replace(' ', '_')}_{op_name}_{s['fh_per_year']}FH"

    # ── EXCEL EXPORT ──
    st.markdown(f'<div class="sec-head">{svg_icon("file", 20)} <span>Excel</span> Report</div>', unsafe_allow_html=True)

    def generate_excel():
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # --- Sheet 1: Summary ---
        ws1 = wb.active
        ws1.title = "DMC Summary"

        # Colors
        navy_fill = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
        blue_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_font = Font(name="Arial", color="FFFFFF", bold=True)
        blue_font = Font(name="Arial", color="2563EB", bold=True, size=11)
        title_font = Font(name="Arial", color="FFFFFF", bold=True, size=16)
        sub_font = Font(name="Arial", color="94A3B8", size=10)
        normal_font = Font(name="Arial", size=10)
        bold_font = Font(name="Arial", size=10, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Title banner
        for col in range(1, 9):
            ws1.cell(row=1, column=col).fill = navy_fill
            ws1.cell(row=2, column=col).fill = navy_fill
            ws1.cell(row=3, column=col).fill = navy_fill
        ws1.merge_cells("A1:H1")
        ws1.cell(row=1, column=1, value="MaintEdge -- DMC Report").font = title_font
        ws1.merge_cells("A2:H2")
        ws1.cell(row=2, column=1, value="Deutsche Aircraft GmbH -- Direct Maintenance Cost Calculator").font = sub_font
        ws1.merge_cells("A3:H3")
        from datetime import datetime
        ws1.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = sub_font

        # Parameters section
        row = 5
        ws1.cell(row=row, column=1, value="OPERATIONAL PARAMETERS").font = blue_font
        row = 6
        param_data = [
            ("Aircraft Type", s["aircraft_type"]),
            ("Operator", s["operator"] or "N/A"),
            ("Base Country", s["base_country"] or "N/A"),
            ("FH / Year", f"{s['fh_per_year']:,}"),
            ("FC / Year", f"{s['fc_per_year']:,}"),
            ("APU Hrs / Year", f"{s['apu_hrs_per_year']:,}"),
            ("FH/FC Ratio", f"{s['fh_fc_ratio']:.2f}"),
            ("Labour Rate (EUR/hr)", f"{s['labour_rate']:.2f}"),
            ("Environment", active_envs_str),
            ("Blended Env Factor", f"x{blended_env:.3f}"),
            ("Gravel Operations", f"{s['gravel_pct']}% (x{gf:.2f})"),
            ("STOL Operations", f"{s['stol_pct']}% (x{sf:.2f})"),
            ("Factors", "Category-specific (see detail)"),
        ]
        for param, val in param_data:
            ws1.cell(row=row, column=1, value=param).font = bold_font
            ws1.cell(row=row, column=3, value=val).font = normal_font
            if row % 2 == 0:
                for c in range(1, 5):
                    ws1.cell(row=row, column=c).fill = alt_fill
            row += 1

        # Totals
        row += 1
        ws1.cell(row=row, column=1, value="DMC RESULTS").font = blue_font
        row += 1
        unsched_xl = total_dmc * 0.40
        logistics_xl = total_dmc * 0.15
        total_all_in_xl = total_dmc + unsched_xl + logistics_xl
        for label, val in [("Scheduled DMC (EUR/FH)", total_dmc), ("Labour DMC (EUR/FH)", total_labour),
                           ("Material DMC (EUR/FH)", total_material), ("Annual Scheduled DMC (EUR)", total_dmc * s["fh_per_year"]),
                           ("Unscheduled Maintenance (40%)", unsched_xl), ("Logistics & Customs (15%)", logistics_xl),
                           ("All-In DMC Estimate (EUR/FH)", total_all_in_xl), ("All-In Annual (EUR)", total_all_in_xl * s["fh_per_year"])]:
            ws1.cell(row=row, column=1, value=label).font = bold_font
            cell = ws1.cell(row=row, column=3, value=round(val, 2))
            cell.font = Font(name="Arial", size=11, bold=True, color="2563EB")
            cell.number_format = '#,##0.00'
            row += 1

        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 5
        ws1.column_dimensions["C"].width = 25
        ws1.column_dimensions["D"].width = 15

        # --- Sheet 2: Category Summary ---
        ws2 = wb.create_sheet("Category Summary")
        cat_sum = df.groupby("Category").agg({
            "DMC Labour (EUR/FH)": "sum", "DMC Material (EUR/FH)": "sum", "DMC Total (EUR/FH)": "sum"
        }).reset_index()
        cat_sum = cat_sum.sort_values("DMC Total (EUR/FH)", ascending=False)
        cat_sum["% of Total"] = (cat_sum["DMC Total (EUR/FH)"] / total_dmc * 100).round(1)

        headers2 = ["Category", "Labour (EUR/FH)", "Material (EUR/FH)", "Total (EUR/FH)", "% of Total"]
        for c, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r_idx, row_data in enumerate(cat_sum.itertuples(index=False), 2):
            ws2.cell(row=r_idx, column=1, value=row_data[0]).font = normal_font
            for c in range(1, 4):
                cell = ws2.cell(row=r_idx, column=c+1, value=round(row_data[c], 2))
                cell.font = normal_font
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="center")
            ws2.cell(row=r_idx, column=5, value=round(row_data[4], 1)).number_format = '0.0'
            if r_idx % 2 == 0:
                for c in range(1, 6):
                    ws2.cell(row=r_idx, column=c).fill = alt_fill
        for c in range(1, 6):
            ws2.column_dimensions[get_column_letter(c)].width = 22

        # --- Sheet 3: Full Detail ---
        ws3 = wb.create_sheet("Detailed Breakdown")
        detail_cols = ["Category", "Inspection", "Interval 1", "Interval 2", "MH",
                       "Material (EUR)", "Occ/yr (Used)", "Driver",
                       "DMC Labour (EUR/FH)", "DMC Material (EUR/FH)", "DMC Total (EUR/FH)"]
        for c, h in enumerate(detail_cols, 1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for c, col_name in enumerate(detail_cols, 1):
                val = row_data[col_name]
                cell = ws3.cell(row=r_idx, column=c, value=val)
                cell.font = normal_font
                if isinstance(val, float):
                    cell.number_format = '#,##0.0000' if "EUR/FH" in col_name else '#,##0.00'
                cell.alignment = Alignment(horizontal="center") if c > 2 else Alignment()
            if r_idx % 2 == 0:
                for c in range(1, len(detail_cols) + 1):
                    ws3.cell(row=r_idx, column=c).fill = alt_fill
        for c in range(1, len(detail_cols) + 1):
            ws3.column_dimensions[get_column_letter(c)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    excel_data = generate_excel()
    st.download_button(
        label="Download Excel Report (.xlsx)",
        data=excel_data,
        file_name=f"{base_filename}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ── PDF EXPORT ──
    st.markdown(f'<div class="sec-head">{svg_icon("file", 20)} <span>PDF</span> Report</div>', unsafe_allow_html=True)

    def generate_pdf():
        import io
        import tempfile
        import os
        from datetime import datetime
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, HRFlowable, Image, KeepTogether)

        buffer = io.BytesIO()
        NAVY       = colors.HexColor("#0A1628")
        DARK_BLUE  = colors.HexColor("#1E293B")
        BRAND_BLUE = colors.HexColor("#2563EB")
        LIGHT_BLUE = colors.HexColor("#EFF6FF")
        SLATE_600  = colors.HexColor("#475569")
        SLATE_400  = colors.HexColor("#94A3B8")
        SLATE_200  = colors.HexColor("#E2E8F0")
        SLATE_50   = colors.HexColor("#F8FAFC")
        WHITE      = colors.white
        GREEN      = colors.HexColor("#22C55E")
        CYAN_400   = colors.HexColor("#60A5FA")
        page_w, page_h = A4

        # ── Chart data ───────────────────────────────────────────────
        cat_sum_pdf = df.groupby("Category").agg({
            "DMC Labour (EUR/FH)": "sum", "DMC Material (EUR/FH)": "sum", "DMC Total (EUR/FH)": "sum"
        }).reset_index().sort_values("DMC Total (EUR/FH)", ascending=False)
        cat_sum_pdf["pct"] = (cat_sum_pdf["DMC Total (EUR/FH)"] / total_dmc * 100).round(1)
        CHART_COLORS = ["#2563EB","#0891B2","#059669","#D97706","#7C3AED",
                        "#DC2626","#475569","#EC4899","#F59E0B","#10B981","#6366F1"]

        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        import numpy as np

        # Pie chart (donut) using matplotlib -- no kaleido/Chrome needed
        fig_pie_mpl, ax_pie = plt.subplots(figsize=(8, 5))
        wedge_colors = CHART_COLORS[:len(cat_sum_pdf)]
        wedges, _, autotexts = ax_pie.pie(
            cat_sum_pdf["DMC Total (EUR/FH)"].round(2),
            labels=None,
            colors=wedge_colors,
            autopct="%1.1f%%",
            pctdistance=0.75,
            startangle=90,
            wedgeprops=dict(width=0.55),
        )
        for at in autotexts:
            at.set_fontsize(8); at.set_color("white")
        legend_patches = [mpatches.Patch(color=wedge_colors[i], label=cat_sum_pdf["Category"].iloc[i])
                          for i in range(len(cat_sum_pdf))]
        ax_pie.legend(handles=legend_patches, loc="center left", bbox_to_anchor=(1, 0.5), fontsize=8)
        ax_pie.set_title("DMC Distribution by Category", fontsize=13, color="#1E293B", pad=15)
        fig_pie_mpl.patch.set_facecolor("white")
        pie_buf = io.BytesIO()
        fig_pie_mpl.savefig(pie_buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig_pie_mpl)
        pie_buf.seek(0)
        pie_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        pie_file.write(pie_buf.read()); pie_file.close()

        # Bar chart (stacked) using matplotlib
        fig_bar_mpl, ax_bar = plt.subplots(figsize=(8, 5))
        x = np.arange(len(cat_sum_pdf))
        labour = cat_sum_pdf["DMC Labour (EUR/FH)"].round(2).values
        material = cat_sum_pdf["DMC Material (EUR/FH)"].round(2).values
        ax_bar.bar(x, labour, 0.6, label="Labour", color="#2563EB")
        ax_bar.bar(x, material, 0.6, bottom=labour, label="Material", color="#22D3EE")
        ax_bar.set_xticks(x)
        ax_bar.set_xticklabels(cat_sum_pdf["Category"], rotation=-40, ha="left", fontsize=8)
        ax_bar.set_ylabel("EUR / FH", fontsize=10)
        ax_bar.set_title("Labour vs Material by Category (EUR/FH)", fontsize=13, color="#1E293B", pad=15)
        ax_bar.legend(loc="upper right", fontsize=9)
        ax_bar.set_facecolor("#F8FAFC")
        ax_bar.grid(axis="y", color="#E2E8F0", linewidth=0.5)
        fig_bar_mpl.patch.set_facecolor("white")
        bar_buf = io.BytesIO()
        fig_bar_mpl.savefig(bar_buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig_bar_mpl)
        bar_buf.seek(0)
        bar_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        bar_file.write(bar_buf.read()); bar_file.close()

        # ── Page template ────────────────────────────────────────────
        class PDFTemplate:
            def on_page(self, canvas, doc):
                canvas.saveState()
                canvas.setStrokeColor(BRAND_BLUE); canvas.setLineWidth(1.5)
                canvas.line(15*mm, page_h-12*mm, page_w-15*mm, page_h-12*mm)
                canvas.setFont("Helvetica-Bold", 7); canvas.setFillColor(SLATE_400)
                canvas.drawString(15*mm, page_h-10*mm, "MAINTEDGE  |  DIRECT MAINTENANCE COST REPORT")
                canvas.drawRightString(page_w-15*mm, page_h-10*mm, "DEUTSCHE AIRCRAFT GMBH")
                canvas.setStrokeColor(SLATE_200); canvas.setLineWidth(0.5)
                canvas.line(15*mm, 14*mm, page_w-15*mm, 14*mm)
                canvas.setFont("Helvetica", 6.5); canvas.setFillColor(SLATE_400)
                canvas.drawString(15*mm, 10*mm, "CONFIDENTIAL  --  Deutsche Aircraft GmbH  --  For authorized use only")
                canvas.drawRightString(page_w-15*mm, 10*mm, f"Page {doc.page}")
                canvas.restoreState()

            def on_first_page(self, canvas, doc):
                canvas.saveState()
                canvas.setFillColor(NAVY); canvas.rect(0, page_h-52*mm, page_w, 52*mm, fill=1, stroke=0)
                canvas.setStrokeColor(BRAND_BLUE); canvas.setLineWidth(2.5)
                canvas.line(0, page_h-52*mm, page_w, page_h-52*mm)
                canvas.setFont("Helvetica", 7); canvas.setFillColor(SLATE_400)
                canvas.drawString(20*mm, page_h-14*mm, "DEUTSCHE AIRCRAFT GMBH")
                canvas.setFont("Helvetica-Bold", 24); canvas.setFillColor(WHITE)
                canvas.drawString(20*mm, page_h-26*mm, "MaintEdge")
                canvas.setFont("Helvetica", 12); canvas.setFillColor(colors.HexColor("#CBD5E1"))
                canvas.drawString(20*mm, page_h-34*mm, "Direct Maintenance Cost Report")
                mod_str_hdr = f"  |  {s.get('mod_variant','')}" if s.get('mod_variant','N/A') != 'N/A' else ""
                canvas.setFont("Helvetica", 8); canvas.setFillColor(SLATE_400)
                canvas.drawString(20*mm, page_h-43*mm, f"{s['aircraft_type']}{mod_str_hdr}  |  {s['operator'] or 'N/A'}  |  {s['base_country'] or 'N/A'}  |  {datetime.now().strftime('%d %B %Y')}")
                # DMC box top-right (green border added)
                bx = page_w-68*mm; by = page_h-46*mm
                canvas.setFillColor(colors.HexColor("#162240"))
                canvas.roundRect(bx, by, 50*mm, 26*mm, 3*mm, fill=1, stroke=0)
                canvas.setStrokeColor(GREEN); canvas.setLineWidth(1.2)
                canvas.roundRect(bx, by, 50*mm, 26*mm, 3*mm, fill=0, stroke=1)
                canvas.setFont("Helvetica", 6.5); canvas.setFillColor(CYAN_400)
                canvas.drawCentredString(bx+25*mm, by+19*mm, "TOTAL DMC")
                canvas.setFont("Helvetica-Bold", 15); canvas.setFillColor(WHITE)
                canvas.drawCentredString(bx+25*mm, by+10*mm, f"EUR {total_dmc:,.2f}")
                canvas.setFont("Helvetica", 7); canvas.setFillColor(SLATE_400)
                canvas.drawCentredString(bx+25*mm, by+4*mm, "per Flight Hour")
                canvas.setStrokeColor(SLATE_200); canvas.setLineWidth(0.5)
                canvas.line(15*mm, 14*mm, page_w-15*mm, 14*mm)
                canvas.setFont("Helvetica", 6.5); canvas.setFillColor(SLATE_400)
                canvas.drawString(15*mm, 10*mm, "CONFIDENTIAL  --  Deutsche Aircraft GmbH  --  For authorized use only")
                canvas.drawRightString(page_w-15*mm, 10*mm, f"Page {doc.page}")
                canvas.restoreState()

        tmpl = PDFTemplate()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
            topMargin=56*mm, bottomMargin=18*mm, leftMargin=15*mm, rightMargin=15*mm)

        # ── Styles ───────────────────────────────────────────────────
        # keepWithNext=1 ensures section heading always stays on same page as its content
        SH  = ParagraphStyle("SH",  fontName="Helvetica-Bold", fontSize=11, textColor=NAVY,
                             spaceBefore=8, spaceAfter=2, leading=14, keepWithNext=1)
        B9  = ParagraphStyle("B9",  fontName="Helvetica",      fontSize=8.5, textColor=SLATE_600,
                             spaceAfter=2, leading=12, alignment=TA_JUSTIFY)
        SM  = ParagraphStyle("SM",  fontName="Helvetica",      fontSize=7,   textColor=SLATE_400,
                             spaceAfter=1, leading=9)
        # Cell styles for detail table -- Paragraph wrapping prevents overflow
        CS  = ParagraphStyle("CS",  fontName="Helvetica",      fontSize=5.5, textColor=SLATE_600,
                             leading=7,   wordWrap="LTR")
        CSH = ParagraphStyle("CSH", fontName="Helvetica-Bold", fontSize=5.8, textColor=WHITE,
                             leading=7.5, alignment=TA_CENTER)
        def hr():   return HRFlowable(width="100%", thickness=0.4, color=SLATE_200, spaceAfter=3)
        def sp(n):  return Spacer(1, n*mm)

        story = []
        labour_pct_val   = (total_labour / total_dmc * 100) if total_dmc > 0 else 0
        material_pct_val = 100 - labour_pct_val
        unsched      = total_dmc * 0.40
        logistics    = total_dmc * 0.15
        total_all_in = total_dmc + unsched + logistics
        mod_str = s.get('mod_variant', 'N/A')
        active_envs_display = ", ".join([f"{e.split('/')[0].strip()} {p}%" for e, p in s.get("env_mix", {}).items() if p > 0])

        # ── EXECUTIVE SUMMARY (KeepTogether: heading + hr + first paragraph) ──
        story.append(KeepTogether([
            Paragraph("Executive Summary", SH), hr(),
            Paragraph(
                f"This report presents the Direct Maintenance Cost (DMC) analysis for the <b>{s['aircraft_type']}"
                f"{' (' + mod_str + ')' if mod_str != 'N/A' else ''}</b> operated by "
                f"<b>{s['operator'] or 'N/A'}</b> based in <b>{s['base_country'] or 'N/A'}</b>. "
                f"The analysis is based on an annual utilization of <b>{s['fh_per_year']:,} flight hours</b> and "
                f"<b>{s['fc_per_year']:,} flight cycles</b>, with a labour rate of <b>EUR {s['labour_rate']:.2f}/hr</b>.",
                B9),
        ]))
        story.append(Paragraph(
            f"The computed <b>scheduled DMC is EUR {total_dmc:,.2f} per flight hour</b> "
            f"(EUR {total_dmc * s['fh_per_year']:,.0f} per year), comprising EUR {total_labour:,.2f}/FH labour "
            f"({labour_pct_val:.0f}%) and EUR {total_material:,.2f}/FH material ({material_pct_val:.0f}%). "
            f"Including estimated unscheduled maintenance (40%) and logistics costs (15%), the "
            f"<b>all-in DMC estimate is EUR {total_all_in:,.2f} per flight hour</b> "
            f"(EUR {total_all_in * s['fh_per_year']:,.0f} per year).",
            B9))

        # Top 3 cost drivers
        top3 = cat_sum_pdf.head(3)
        drivers = ", ".join([f"{r['Category']} ({r['pct']:.0f}%)" for _, r in top3.iterrows()])
        story.append(Paragraph(
            f"The dominant cost drivers are: <b>{drivers}</b>. "
            f"Environmental profile: {active_envs_display}. "
            f"Gravel operations: {s['gravel_pct']}%, STOL: {s['stol_pct']}%"
            f"{', High Altitude: ' + str(s.get('env_mix',{}).get('High Altitude',0)) + '%' if s.get('env_mix',{}).get('High Altitude',0) > 0 else ''}.",
            B9))
        story.append(sp(2))

        # ── TOTAL DMC GREEN BOX (matches web UI highlighted card) ────
        # Four metric cards in a row: Total DMC (dark+green border) | Labour | Material | All-In
        _lbl = lambda t, c: Paragraph(t, ParagraphStyle("_l", fontName="Helvetica-Bold",
            fontSize=7, textColor=c, leading=9, alignment=TA_CENTER))
        _val = lambda t, c, sz=11: Paragraph(t, ParagraphStyle("_v", fontName="Helvetica-Bold",
            fontSize=sz, textColor=c, leading=sz+2, alignment=TA_CENTER))
        _sub = lambda t: Paragraph(t, ParagraphStyle("_s", fontName="Helvetica",
            fontSize=6.5, textColor=SLATE_400, leading=8, alignment=TA_CENTER))

        metrics_tbl = Table([
            [_lbl("TOTAL DMC", CYAN_400),      _lbl("Labour DMC", SLATE_600),           _lbl("Material DMC", SLATE_600),         _lbl("All-In Estimate", SLATE_600)],
            [_val(f"EUR {total_dmc:,.2f}", WHITE, 12), _val(f"EUR {total_labour:,.2f}", BRAND_BLUE, 10), _val(f"EUR {total_material:,.2f}", BRAND_BLUE, 10), _val(f"EUR {total_all_in:,.2f}", BRAND_BLUE, 10)],
            [_sub("per Flight Hour"),           _sub("per Flight Hour"),                 _sub("per Flight Hour"),                 _sub("incl. 40% + 15%")],
        ], colWidths=[47*mm, 43*mm, 43*mm, 43*mm])
        metrics_tbl.setStyle(TableStyle([
            # First column: dark navy background + green border
            ("BACKGROUND",    (0,0), (0,-1), NAVY),
            ("BOX",           (0,0), (0,-1), 1.5, GREEN),
            # Other columns: light background
            ("BACKGROUND",    (1,0), (-1,-1), SLATE_50),
            ("BOX",           (1,0), (1,-1), 0.4, SLATE_200),
            ("BOX",           (2,0), (2,-1), 0.4, SLATE_200),
            ("BOX",           (3,0), (3,-1), 0.4, SLATE_200),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        story.append(metrics_tbl)
        story.append(sp(4))

        # ── 1. PARAMETERS ──
        param_rows = [
            ["Parameter", "Value", "Parameter", "Value"],
            ["Aircraft", s["aircraft_type"], "MOD Variant", mod_str],
            ["Operator", s["operator"] or "N/A", "Labour Rate", f"EUR {s['labour_rate']:.2f}/hr"],
            ["Base Country", s["base_country"] or "N/A", "FH/FC Ratio", f"{s['fh_fc_ratio']:.2f}"],
            ["FH / Year", f"{s['fh_per_year']:,}", "FC / Year", f"{s['fc_per_year']:,}"],
            ["APU Hrs / Year", "N/A (no APU)" if "eco" in s["aircraft_type"].lower() else f"{s['apu_hrs_per_year']:,}", "Environment", active_envs_display[:40]],
            ["Gravel Ops", f"{s['gravel_pct']}%", "STOL Ops", f"{s['stol_pct']}%"],
        ]
        pt = Table(param_rows, colWidths=[32*mm, 48*mm, 32*mm, 48*mm])
        pt.setStyle(TableStyle([
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("BACKGROUND",(0,0),(-1,0),DARK_BLUE),("TEXTCOLOR",(0,0),(-1,0),WHITE),
            ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,1),(2,-1),"Helvetica-Bold"),
            ("TEXTCOLOR",(0,1),(0,-1),DARK_BLUE),("TEXTCOLOR",(2,1),(2,-1),DARK_BLUE),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,SLATE_50]),
            ("GRID",(0,0),(-1,-1),0.3,SLATE_200),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),("LEFTPADDING",(0,0),(-1,-1),5),
        ]))
        story.append(KeepTogether([Paragraph("1. Operational Parameters", SH), hr(), pt]))
        story.append(sp(3))

        # ── 2. DMC SUMMARY ──
        sum_rows = [
            ["Component", "EUR / FH", "EUR / Year", "%"],
            ["Scheduled DMC", f"{total_dmc:,.2f}", f"{total_dmc * s['fh_per_year']:,.0f}", "100%"],
            ["  Labour", f"{total_labour:,.2f}", f"{total_labour * s['fh_per_year']:,.0f}", f"{labour_pct_val:.0f}%"],
            ["  Material", f"{total_material:,.2f}", f"{total_material * s['fh_per_year']:,.0f}", f"{material_pct_val:.0f}%"],
            ["Unscheduled (40%)", f"{unsched:,.2f}", f"{unsched * s['fh_per_year']:,.0f}", "40%"],
            ["Logistics (15%)", f"{logistics:,.2f}", f"{logistics * s['fh_per_year']:,.0f}", "15%"],
            ["All-In Estimate", f"{total_all_in:,.2f}", f"{total_all_in * s['fh_per_year']:,.0f}", "155%"],
        ]
        st2 = Table(sum_rows, colWidths=[45*mm, 35*mm, 40*mm, 20*mm])
        st2.setStyle(TableStyle([
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),WHITE),
            ("BACKGROUND",(0,1),(-1,1),LIGHT_BLUE),("FONTNAME",(0,1),(-1,1),"Helvetica-Bold"),
            ("TEXTCOLOR",(1,1),(1,1),BRAND_BLUE),("FONTSIZE",(1,1),(1,1),10),
            ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
            ("BACKGROUND",(0,-1),(-1,-1),NAVY),("TEXTCOLOR",(0,-1),(-1,-1),WHITE),("FONTSIZE",(0,-1),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,2),(-1,-2),[WHITE,SLATE_50]),
            ("GRID",(0,0),(-1,-1),0.3,SLATE_200),("ALIGN",(1,0),(-1,-1),"CENTER"),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("LEFTPADDING",(0,0),(0,-1),6),
        ]))
        story.append(KeepTogether([Paragraph("2. DMC Summary", SH), hr(), st2]))
        story.append(sp(3))

        # ── 3. CATEGORY BREAKDOWN WITH CHARTS ──
        # Charts stacked full-width for maximum readability
        pie_img = Image(pie_file.name, width=168*mm, height=105*mm)
        bar_img = Image(bar_file.name, width=168*mm, height=105*mm)
        cat_subtitle = Paragraph(f"{len(cat_sum_pdf)} cost categories shown.", SM)
        story.append(KeepTogether([Paragraph("3. Category Breakdown", SH), hr(), cat_subtitle, sp(1), pie_img, sp(2), bar_img]))
        story.append(sp(3))

        # Category summary table
        cat_rows = [["#", "Category", "Labour (EUR/FH)", "Material (EUR/FH)", "Total (EUR/FH)", "%"]]
        for idx, (_, r) in enumerate(cat_sum_pdf.iterrows(), 1):
            cat_rows.append([str(idx), r["Category"], f"{r['DMC Labour (EUR/FH)']:.2f}",
                f"{r['DMC Material (EUR/FH)']:.2f}", f"{r['DMC Total (EUR/FH)']:.2f}", f"{r['pct']:.1f}%"])
        ct = Table(cat_rows, colWidths=[8*mm, 60*mm, 28*mm, 28*mm, 28*mm, 16*mm])
        ct.setStyle(TableStyle([
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("BACKGROUND",(0,0),(-1,0),DARK_BLUE),("TEXTCOLOR",(0,0),(-1,0),WHITE),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,SLATE_50]),("GRID",(0,0),(-1,-1),0.3,SLATE_200),
            ("ALIGN",(0,0),(0,-1),"CENTER"),("ALIGN",(2,0),(-1,-1),"CENTER"),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("LEFTPADDING",(1,0),(1,-1),5),
        ]))
        story.append(KeepTogether([Paragraph("Category Summary", ParagraphStyle("SH2",
            fontName="Helvetica-Bold", fontSize=9, textColor=DARK_BLUE, spaceBefore=4,
            spaceAfter=2, leading=12, keepWithNext=1)), ct]))
        story.append(sp(3))

        # ── 4. DETAILED BREAKDOWN ──
        # All cells wrapped in Paragraph objects so long text wraps properly (fixes PWC row overflow)
        det_col_hdrs = ["Inspection", "Category", "Int 1", "Int 2", "MH", "Mat (EUR)", "Occ/yr", "Adj.F", "EUR/FH"]
        det_col_w    = [40*mm, 20*mm, 16*mm, 16*mm, 10*mm, 20*mm, 14*mm, 11*mm, 21*mm]  # total = 168mm
        det_rows = [[Paragraph(h, CSH) for h in det_col_hdrs]]
        for _, r in df.iterrows():
            cat_short = (r["Category"]
                .replace("Corrosion Prevention ", "CP ")
                .replace("Structural Sampling ",  "SSI ")
                .replace("Fatigue Damage ",        "FD ")
                .replace("Time Controlled ",       "TC ")
                .replace("Engines (PWC FMP)",      "Engines\n(PWC FMP)"))
            det_rows.append([
                Paragraph(r["Inspection"],               CS),
                Paragraph(cat_short,                     CS),
                Paragraph(str(r["Interval 1"]),          CS),
                Paragraph(str(r["Interval 2"]),          CS),
                Paragraph(str(r["MH"]),                  CS),
                Paragraph(f"{r['Material (EUR)']:,.0f}", CS),
                Paragraph(f"{r['Occ/yr (Used)']:.3f}",  CS),
                Paragraph(f"{r['Adj. Factor']:.3f}",     CS),
                Paragraph(f"{r['DMC Total (EUR/FH)']:.4f}", CS),
            ])
        dt = Table(det_rows, colWidths=det_col_w, repeatRows=1)
        dt.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  DARK_BLUE),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, SLATE_50]),
            ("GRID",          (0,0), (-1,-1), 0.2, SLATE_200),
            ("ALIGN",         (0,0), (0,-1),  "LEFT"),
            ("ALIGN",         (1,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
            ("TOPPADDING",    (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("LEFTPADDING",   (0,0), (0,-1),  3),
        ]))
        detail_subtitle = Paragraph(f"{len(df)} maintenance items. Category-specific adjustment factors applied.", SM)
        story.append(KeepTogether([Paragraph("4. Detailed Item Breakdown", SH), hr(), detail_subtitle, sp(1), dt]))
        story.append(sp(3))

        # ── DISCLAIMER ──
        story.append(HRFlowable(width="100%", thickness=0.4, color=SLATE_200, spaceAfter=2))
        story.append(Paragraph("Disclaimer & Notes", ParagraphStyle("DH", fontName="Helvetica-Bold", fontSize=7.5, textColor=DARK_BLUE, spaceAfter=1)))
        disc = ParagraphStyle("D", fontName="Helvetica", fontSize=6.5, textColor=SLATE_600, leading=8.5, spaceAfter=1, alignment=TA_JUSTIFY)
        story.append(Paragraph("1. This report is generated by MaintEdge v1.0 by Deutsche Aircraft GmbH. All cost figures are estimates based on the OEM maintenance programme, assumed utilization profiles, and engineering-derived adjustment factors. Actual costs may vary depending on operator practices, MRO capabilities, and regulatory requirements.", disc))
        story.append(Paragraph("2. Unscheduled maintenance (40%) and logistics (15%) are industry benchmarks for regional turboprops and should be validated against operator experience.", disc))
        story.append(Paragraph("3. Environmental and operational factors are category-specific, based on engineering judgement and published industry data. Users should apply professional judgement for contractual or financial planning.", disc))
        story.append(Paragraph("4. Material costs reflect current OEM pricing, subject to escalation, discount agreements, and exchange rate fluctuations. Labour rates are user-defined.", disc))
        story.append(Paragraph("5. This document does not constitute a binding cost commitment or warranty by Deutsche Aircraft GmbH.", disc))
        story.append(sp(2))
        story.append(Paragraph("MaintEdge by Deutsche Aircraft GmbH  |  Confidential  |  For authorized use only",
            ParagraphStyle("FN", fontName="Helvetica-Bold", fontSize=6.5, textColor=SLATE_400, alignment=TA_CENTER)))

        doc.build(story, onFirstPage=tmpl.on_first_page, onLaterPages=tmpl.on_page)

        # Cleanup temp files
        import os
        try: os.unlink(pie_file.name)
        except: pass
        try: os.unlink(bar_file.name)
        except: pass

        return buffer.getvalue()

    try:
        pdf_data = generate_pdf()
        st.download_button(
            label="Download PDF Report (.pdf)",
            data=pdf_data,
            file_name=f"{base_filename}.pdf",
            mime="application/pdf",
        )
    except ImportError:
        st.warning("PDF export requires reportlab. Install with: pip install reportlab")

    # Parameters reference
    st.markdown("---")
    st.markdown(f'<div class="sec-head">{svg_icon("settings", 20)} <span>Report</span> Parameters</div>', unsafe_allow_html=True)

    params = {
        "Parameter": [
            "Aircraft Type", "MOD Variant", "Operator", "Base Country", "FH/Year", "FC/Year",
            "APU Hrs/Year", "FH/FC Ratio", "Labour Rate", "Environment Mix",
            "Blended Env Factor", "Gravel %", "Gravel Factor", "STOL %", "STOL Factor",
            "Factors", "Category-Specific",
        ],
        "Value": [
            s["aircraft_type"], s.get("mod_variant", "N/A"), s["operator"] or "N/A", s["base_country"] or "N/A",
            f"{s['fh_per_year']:,}", f"{s['fc_per_year']:,}", f"{s['apu_hrs_per_year']:,}",
            f"{s['fh_fc_ratio']:.2f}", f"EUR {s['labour_rate']:.2f}/hr", active_envs_str,
            f"x{blended_env:.3f}", f"{s['gravel_pct']}%", f"x{gf:.2f}", f"{s['stol_pct']}%", f"x{sf:.2f}",
            "Env x Gravel x STOL", "Weighted per category (see Calculate page)",
        ],
    }
    st.dataframe(pd.DataFrame(params), use_container_width=True, hide_index=True)


# Footer
st.markdown('<div class="footer"><strong>MaintEdge</strong> by Deutsche Aircraft GmbH -- DMC Calculator v1.0.0<br>Confidential -- For authorized use only</div>', unsafe_allow_html=True)
